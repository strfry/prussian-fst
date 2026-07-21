#!/usr/bin/env python3
"""Delta-Review: Korpus ↔ Twanksta-Dict — mit Occurrence-Tracking,
Youtubebe-Links, Eigennamen-/Tippfehler-/Interjektions-/Lehnwort-Erkennung
und Error-Group-Mapping.

Pipeline:
  1. Precondition: verify dict is full-form
  2. Tokenize corpus → TypeEntry-Liste mit Occurrences und Source-Refs
  3. Exact lookup each type → discard known
  4. Fuzzy lookup unknowns → classify as signature (A) or OOV (B)
  5. Analysis passes: proper names, typos, interjections,
     suspicious POS matches, loanword endings
  6. Cluster A by transformation signature
  7. Output delta_signatures.xlsx + delta_oov.xlsx +
     delta_error_groups.xlsx + delta_broken_lemmas.xlsx + run report
"""

import json
import re
import subprocess
import unicodedata
import sys
import time
from collections import defaultdict, Counter
from dataclasses import dataclass, field
from pathlib import Path

try:
    from rapidfuzz.distance import Levenshtein as _Lev
    from rapidfuzz import process as _rf_process
    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

REPO = Path(__file__).resolve().parents[2]
TWANKSTA_PATH = REPO / "../prussian-corpus/parsed/twanksta_entries.json"
CORPUS_PATH = REPO / "../prussian-corpus/parsed/youtube_corpus_sentences.json"
OUT_DIR = REPO / "data/reports"
OUT_SIG = OUT_DIR / "delta_signatures.xlsx"
OUT_OOV = OUT_DIR / "delta_oov.xlsx"
OUT_BROKEN = OUT_DIR / "delta_broken_lemmas.xlsx"
OUT_ERR_GROUPS = OUT_DIR / "delta_error_groups.xlsx"

FST_DIR = Path(__file__).resolve().parents[2]
BASE_FST = FST_DIR / "build/base.fst"
LENIENT_FST = FST_DIR / "build/lenient.fst"

MATCH_POLICY = ("ED<=1 beliebig, OP-Plaublitaets-Ranking "
                "(Gemination > Vokal > j > Sub > Ins/Del); "
                "ED=2 nur plausible Ops (Gemination/Vokal/j)")
SKIP_VIDEOS = {"qLwBCWtMuH8"}


@dataclass
class SourceRef:
    video_id: str
    title: str
    start: str   # "00:02:19,300"
    end: str

    @property
    def t_seconds(self) -> int:
        try:
            clean = self.start.replace(",", ".")
            h, m, s = clean.split(":")
            return int(h) * 3600 + int(m) * 60 + int(float(s))
        except (ValueError, AttributeError):
            return 0

    @property
    def youtube_url(self) -> str:
        ts = self.t_seconds
        return f"https://youtu.be/{self.video_id}?t={ts}" if ts else f"https://youtu.be/{self.video_id}"


@dataclass
class Occurrence:
    context: str       # full sentence text
    position: int      # 0-based index of this token in sentence
    position_type: str  # "sentence_initial", "sentence_medial", "only_token"
    form_with_case: str  # original casing as it appeared here
    sentence_frequency: int
    sources: list = field(default_factory=list)  # list of SourceRef


@dataclass
class TypeEntry:
    form: str            # canonical form (first encounter, preserves case)
    folded: str
    frequency: int       # total across all sentences × positions
    occurrences: list = field(default_factory=list)  # list of Occurrence
    dict_lemma: str = ""
    is_proper_name: bool = False
    is_likely_typo: bool = False
    corpus_freq_of_match: int = 0
    suspicious_pos: str = ""
    interjection: bool = False
    loanword_ending: str = ""

    @property
    def source_links(self) -> str:
        urls = []
        seen = set()
        for occ in self.occurrences:
            for sr in occ.sources:
                key = (sr.video_id, sr.t_seconds)
                if key not in seen:
                    seen.add(key)
                    urls.append(sr.youtube_url)
        return "\n".join(urls)

    @property
    def all_contexts(self) -> str:
        lines = []
        for occ in self.occurrences:
            lines.append(f"[×{occ.sentence_frequency}] {occ.context}")
        return "\n".join(lines)

    def pick_best_occurrence(self):
        """Return the occurrence with most sources (best linked example)."""
        if not self.occurrences:
            return None
        return max(self.occurrences,
                   key=lambda o: len(o.sources) if o.sources else 0)

    def top_links(self, n: int = 3) -> str:
        """First N unique youtubebe links across occurrences."""
        urls = []
        seen = set()
        for occ in self.occurrences:
            for sr in occ.sources:
                key = (sr.video_id, sr.t_seconds)
                if key not in seen:
                    seen.add(key)
                    urls.append(sr.youtube_url)
                    if len(urls) >= n:
                        return "\n".join(urls)
        return "\n".join(urls)

    def example_context(self) -> str:
        """One example sentence for display (prefer one with Youtubebe sources)."""
        best = self.pick_best_occurrence()
        if best and best.sources:
            return f"[×{best.sentence_frequency}] {best.context}"
        if self.occurrences:
            occ = self.occurrences[0]
            return f"[×{occ.sentence_frequency}] {occ.context}"
        return ""


def timestamp_to_seconds(ts: str) -> int:
    try:
        clean = ts.replace(",", ".")
        h, m, s = clean.split(":")
        return int(h) * 3600 + int(m) * 60 + int(float(s))
    except (ValueError, AttributeError):
        return 0


def make_youtube_url(video_id: str, t_seconds: int = 0) -> str:
    if t_seconds:
        return f"https://youtu.be/{video_id}?t={t_seconds}"
    return f"https://youtu.be/{video_id}"


def fold(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return unicodedata.normalize("NFC", s.lower())


def extract_prussian_tokens_with_pos(text: str) -> list[tuple[str, int, int]]:
    """Return (token, position, first_alpha_position) tuples.
    position_type is determined later by comparing position to
    first_alpha_position."""
    if re.match(r'^[A-Z]{2,5}:', text):
        if not text.startswith('PR:'):
            return []
        text = text[3:].lstrip()
    text = text.split("//")[0]
    text = re.sub(r'\[[^\]]*\]', '', text)
    text = re.sub(r'\([^)]*\)', '', text)
    tokens = []
    pos = 0
    first_alpha = None
    for tok in text.split():
        tok = tok.lstrip('=/')
        tok = tok.strip('.,!?;:()[]{}«»""\' \t')
        if tok and len(tok) >= 2 and tok.isalpha():
            if first_alpha is None:
                first_alpha = pos
            tokens.append((tok, pos, first_alpha if first_alpha is not None else 0))
        pos += 1
    return tokens


def edit_distance(a: str, b: str) -> int:
    if HAVE_RAPIDFUZZ:
        return _Lev.distance(a, b)
    m, n = len(a), len(b)
    if m < n:
        a, b = b, a
        m, n = n, m
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[0]
        dp[0] = i
        for j in range(1, n + 1):
            temp = dp[j]
            dp[j] = min(
                prev + (0 if a[i-1] == b[j-1] else 1),
                dp[j] + 1,
                dp[j-1] + 1,
            )
            prev = temp
    return dp[n]


VOWEL_MAP = {'ā': 'a', 'a': 'ā', 'ē': 'e', 'e': 'ē', 'ī': 'i', 'i': 'ī',
             'ō': 'o', 'o': 'ō', 'ū': 'u', 'u': 'ū'}
CONSONANTS = set('bcdfghjklmnprstvwzšž')
VOWELS = set('aeiouāēīōū')
FOLDED_VOWELS = set('aeiou')

W_VL, W_SUB, W_INDEL = 3, 10, 10


def _sub_cost(ca: str, cb: str) -> int:
    if ca == cb:
        return 0
    if VOWEL_MAP.get(ca) == cb:
        return W_VL
    return W_SUB


def edit_ops(a: str, b: str) -> list[tuple[str, str, str, int, int]]:
    m, n = len(a), len(b)
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(m + 1): dp[i][0] = i * W_INDEL
    for j in range(n + 1): dp[0][j] = j * W_INDEL
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            dp[i][j] = min(dp[i-1][j] + W_INDEL, dp[i][j-1] + W_INDEL,
                           dp[i-1][j-1] + _sub_cost(a[i-1], b[j-1]))
    ops = []
    i, j = m, n
    while i > 0 or j > 0:
        if (i > 0 and j > 0
                and dp[i][j] == dp[i-1][j-1] + _sub_cost(a[i-1], b[j-1])):
            op = 'eq' if a[i-1] == b[j-1] else 'sub'
            ops.append((op, a[i-1], b[j-1], i-1, j-1))
            i -= 1; j -= 1
        elif i > 0 and dp[i][j] == dp[i-1][j] + W_INDEL:
            ops.append(('del', a[i-1], '', i-1, j))
            i -= 1
        else:
            ops.append(('ins', '', b[j-1], i, j-1))
            j -= 1
    ops.reverse()
    return ops


def ctx_label(ch: str, boundary: str = '#') -> str:
    if ch == '' or ch is None:
        return boundary
    if ch in CONSONANTS:
        return 'C'
    if ch in VOWELS:
        return 'V'
    return ch if ch.isalpha() else boundary


def derive_signature(ops: list, a: str, b: str) -> str:
    changes = [op for op in ops if op[0] != 'eq']
    if not changes:
        return "EXACT"
    lon = len(a)
    ltg = len(b)
    vlength = [op for op in changes
               if op[0] == 'sub' and VOWEL_MAP.get(op[1]) == op[2]]
    core = [op for op in changes if op not in vlength]
    vl_detail = ",".join(sorted({f"{op[1]}→{op[2]}" for op in vlength}))
    suffix = f" (+VLength {vl_detail})" if vlength else ""
    if not core:
        return f"VLength ({vl_detail})"
    if len(core) == 1:
        op, ca, cb, pa, pb = core[0]
        lctx_a = ctx_label(a[pa-1] if pa > 0 else '')
        rctx_a = ctx_label(a[pa+1] if pa + 1 < lon else '')
        if op == 'sub':
            return f"Sub {ca}→{cb} / {lctx_a}_{rctx_a}{suffix}"
        if op == 'del':
            if (pa > 0 and a[pa-1] == ca) or (pa + 1 < lon and a[pa+1] == ca):
                return f"Gemination {ca}{ca}→{ca} / {lctx_a}_{rctx_a}{suffix}"
            return f"Del {ca} / {lctx_a}_{rctx_a}{suffix}"
        if op == 'ins':
            lctx_b = ctx_label(b[pb-1] if pb > 0 else '')
            rctx_b = ctx_label(b[pb+1] if pb + 1 < ltg else '')
            if (pb > 0 and b[pb-1] == cb) or (pb + 1 < ltg and b[pb+1] == cb):
                return f"Gemination {cb}→{cb}{cb} / {lctx_b}_{rctx_b}{suffix}"
            return f"Ins {cb} / {lctx_b}_{rctx_b}{suffix}"
    if all(op[0] == 'sub' for op in core):
        details = ",".join(f"{op[1]}→{op[2]}" for op in core)
        return f"Sub multi ({details}){suffix}"
    detail = ",".join(f"{op[0]}({op[1] or ''}→{op[2] or ''})" for op in core)
    detail = detail[:60]
    return f"Composite ({detail}){suffix}"


def op_penalty(folded_corpus: str, folded_dict: str) -> int:
    """Penalty fur die Edit-Operation zwischen zwei gefoldeten Formen.
    0=gut (Gemination), 4=schlecht (beliebiges Ins/Del), 5=ED>=2.
    Annahme: edit_distance(folded_corpus, folded_dict) <= 2."""
    a, b = folded_corpus, folded_dict
    la, lb = len(a), len(b)

    if la == lb:
        for i in range(la):
            if a[i] != b[i]:
                if a[i] in FOLDED_VOWELS and b[i] in FOLDED_VOWELS:
                    return 1  # vowel quality
                if a[i] == 'j' or b[i] == 'j':
                    return 2  # j-related
                return 3  # other substitution

    if la + 1 == lb:  # insertion into a
        i = 0
        while i < la and a[i] == b[i]:
            i += 1
        ins = b[i]
        if (i > 0 and b[i-1] == ins) or (i + 1 < lb and b[i+1] == ins):
            return 0  # gemination
        if ins in FOLDED_VOWELS:
            return 2
        if ins == 'j':
            return 2
        return 4  # plain insertion

    if la - 1 == lb:  # deletion from a
        i = 0
        while i < lb and a[i] == b[i]:
            i += 1
        deleted = a[i]
        if (i > 0 and a[i-1] == deleted) or (i + 1 < la and a[i+1] == deleted):
            return 0  # degemination
        if deleted == 'j':
            return 2
        return 4  # plain deletion

    return 5  # ED >= 2


# ── FST-based lookup ──

def fst_lookup_batch(forms: list[str], fst_path: Path) -> dict[str, list[tuple[str, str, list[str]]]]:
    """Batch lookup through hfst-flookup.
    Returns {form: [(lemma, analysis_str, tags), ...]}"""
    if not forms:
        return {}
    proc = subprocess.run(
        ["hfst-flookup", "-q", str(fst_path)],
        input="\n".join(forms) + "\n",
        capture_output=True, text=True, check=True,
    )
    results = defaultdict(list)
    for line in proc.stdout.splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        form, analysis = parts[0], parts[1]
        if analysis.endswith("+?"):
            continue
        segs = analysis.split("+")
        lemma, tags = segs[0], segs[1:]
        if not tags:
            continue
        results[form].append((lemma, analysis, tags))
    return dict(results)


def fst_lookup_types(type_list: list) -> dict[str, tuple[str, str]]:
    """FST-based exact lookup: base.fst → lenient.fst fallback.
    Returns {folded_form: (lemma, analysis_str)} for known types."""
    forms_set = set()
    form_to_folded = {}  # original string → folded

    for te in type_list:
        forms_set.add(te.form)
        form_to_folded[te.form] = te.folded
        # Lowercased variant (for sentence-initial caps)
        lo = te.form.lower()
        if lo != te.form:
            forms_set.add(lo)
            form_to_folded.setdefault(lo, te.folded)
        # All occurrence variants
        for occ in te.occurrences:
            fwc = occ.form_with_case
            if fwc not in forms_set:
                forms_set.add(fwc)
                form_to_folded.setdefault(fwc, te.folded)

    all_forms = sorted(forms_set)

    # 1. base.fst
    print(f"  FST lookup: {len(all_forms)} unique strings on base.fst...", file=sys.stderr)
    known = fst_lookup_batch(all_forms, BASE_FST)
    unknown = [f for f in all_forms if f not in known]

    # 2. lenient.fst fallback
    if unknown:
        print(f"  FST fallback: {len(unknown)} onto lenient.fst...", file=sys.stderr)
        lenient = fst_lookup_batch(unknown, LENIENT_FST)
        for f, analyses in lenient.items():
            known[f] = analyses  # lenient overwrites/preferred

    # Map back: folded_form → best (lemma, analysis)
    result = {}
    for orig_form, analyses in known.items():
        folded = form_to_folded.get(orig_form)
        if folded and folded not in result:
            lemma, analysis, tags = analyses[0]
            result[folded] = (lemma, analysis)

    return result


# ── Dictionary loading ──

def load_dict_forms() -> dict:
    raw = json.loads(TWANKSTA_PATH.read_text(encoding="utf-8"))
    all_forms = set()
    forms_by_len = defaultdict(set)
    folded_to_lemma = defaultdict(list)
    folded_to_orig = {}
    lemma_by_form = {}
    lemma_info = {}  # folded_lemma → {"is_verb": bool, "paradigm": str, "gender": str}

    def _clean(f: str) -> str:
        f = f.strip().strip('!?.,;: ')
        if not f or " " in f or "/" in f:
            return ""
        if not all(c.isalpha() or c == '-' for c in f):
            return ""
        return f

    def _add(f: str, lemma: str):
        f = _clean(f)
        if not f:
            return
        ff = fold(f)
        all_forms.add(ff)
        forms_by_len[len(ff)].add(ff)
        folded_to_lemma[ff].append(lemma)
        folded_to_orig.setdefault(ff, f)
        lemma_by_form.setdefault(f, lemma)

    def _broken_stem(word: str):
        fw = fold(word)
        if " " in fw:
            return None
        is_verb = False
        for suf in ("twei", "tun"):
            if fw.endswith(suf):
                fw = fw[:-len(suf)]
                is_verb = True
                break
        else:
            fw = fw.rstrip('aeiou')
        if len(fw) < 4:
            return None
        return (fw, is_verb)

    broken_lemmas = []

    for e in raw:
        word = e.get("word", "")
        if not word:
            continue
        _add(word, word)

        paradigm = e.get("paradigm", "")
        gender = e.get("gender", "")
        has_ind = bool(e.get("forms", {}).get("indicative"))
        has_opt = bool(e.get("forms", {}).get("optative"))
        has_imp = bool(e.get("forms", {}).get("imperative"))
        is_verb = has_ind or has_opt or has_imp
        if is_verb and not has_ind:
            pass  # some verbs only have non-indicative paradigms

        fl = fold(word)
        lemma_info[fl] = {
            "is_verb": is_verb,
            "paradigm": paradigm,
            "gender": gender,
        }

        decl_cells = []
        for g in e.get("forms", {}).get("declension", []):
            for c in g.get("cases", []):
                for num in ("singular", "plural"):
                    if c.get(num):
                        decl_cells.append(c[num])
        if len(decl_cells) >= 4 and set(decl_cells) == {word}:
            stem_info = _broken_stem(word)
            broken_lemmas.append({
                "word": word,
                "paradigm": paradigm,
                "n_cells": len(decl_cells),
                "stem": stem_info[0] if stem_info else None,
                "is_verb": stem_info[1] if stem_info else False,
            })
        else:
            for f in decl_cells:
                _add(f, word)

        for p in e.get("forms", {}).get("participles", []):
            _add(p.get("form", ""), word)

        for mood in ("indicative", "optative", "imperative", "subjunctive"):
            val = e.get("forms", {}).get(mood)
            if isinstance(val, list):
                for tense_entry in val:
                    if isinstance(tense_entry, dict):
                        for entry in tense_entry.get("forms", []):
                            _add(entry.get("form", ""), word)
                        _add(tense_entry.get("form", ""), word)
            elif isinstance(val, str):
                _add(val, word)

    folded_to_orig_form = {}
    for ff in all_forms:
        orig = folded_to_orig.get(ff)
        if orig:
            folded_to_orig_form[ff] = orig

    forms_by_len = {k: sorted(v) for k, v in forms_by_len.items()}
    n_lemmas = len(raw)

    return {
        "all_forms": all_forms,
        "alphabet": sorted({c for f in all_forms for c in f if c.isalpha()}),
        "forms_by_len": forms_by_len,
        "folded_to_lemma": dict(folded_to_lemma),
        "folded_to_orig": folded_to_orig,
        "folded_to_orig_form": folded_to_orig_form,
        "lemma_by_form": lemma_by_form,
        "n_lemmas": n_lemmas,
        "broken_lemmas": broken_lemmas,
        "lemma_info": lemma_info,
    }


def check_precondition(forms_dict: dict) -> dict:
    lemma_samples = list(forms_dict["lemma_by_form"].items())
    inflected = [(f, lm) for f, lm in lemma_samples if f != lm and len(f) >= 3]
    test = inflected[:10]
    results = []
    for form, lemma in test:
        folded = fold(form)
        hit = folded in forms_dict["folded_to_lemma"]
        results.append({"form": form, "lemma": lemma, "folded": folded, "hit": hit})
    all_hit = all(r["hit"] for r in results)
    return {"tested": results, "all_pass": all_hit,
            "n_entries": len(forms_dict["all_forms"]),
            "n_original_forms": len(forms_dict["lemma_by_form"]),
            "n_lemmas": forms_dict["n_lemmas"]}


# ── Tokenize with occurrence tracking ──

def tokenize_corpus() -> list:
    raw = json.loads(CORPUS_PATH.read_text(encoding="utf-8"))
    type_data = {}  # folded → TypeEntry

    for e in raw:
        sources = e.get("sources", [])
        if sources and all(s.get("video_id") in SKIP_VIDEOS for s in sources):
            continue

        text = e.get("text", "")
        sent_freq = e.get("frequency", 1)

        source_refs = [
            SourceRef(
                video_id=s.get("video_id", ""),
                title=s.get("title", ""),
                start=s.get("start", "00:00:00,000"),
                end=s.get("end", "00:00:00,000"),
            )
            for s in sources
        ]

        tokens = extract_prussian_tokens_with_pos(text)
        for tok, pos, first_alpha in tokens:
            ftok = fold(tok)
            if ftok not in type_data:
                type_data[ftok] = TypeEntry(
                    form=tok, folded=ftok, frequency=0, occurrences=[])

            te = type_data[ftok]
            te.frequency += sent_freq

            if tok[0].isupper():
                pos_type = "sentence_initial" if pos == first_alpha else "sentence_medial"
            else:
                pos_type = "sentence_initial" if pos == first_alpha else "sentence_medial"

            te.occurrences.append(Occurrence(
                context=text,
                position=pos,
                position_type=pos_type,
                form_with_case=tok,
                sentence_frequency=sent_freq,
                sources=source_refs,
            ))

    return sorted(type_data.values(), key=lambda x: -x.frequency)


# ── Lookup ──

def lookup_exact(ftok: str, forms_dict: dict) -> list[str]:
    h = forms_dict["folded_to_lemma"].get(ftok)
    return h if h else []


def ed1_neighbors(w: str, alphabet: list[str]):
    L = len(w)
    for i in range(L):
        yield w[:i] + w[i+1:]
        for ch in alphabet:
            if ch != w[i]:
                yield w[:i] + ch + w[i+1:]
    for i in range(L + 1):
        for ch in alphabet:
            yield w[:i] + ch + w[i:]


def plausible_variants(w: str) -> set[str]:
    out = set()
    L = len(w)
    for i, ch in enumerate(w):
        if i + 1 < L and w[i+1] == ch:
            out.add(w[:i] + w[i+1:])
        if ch not in FOLDED_VOWELS:
            out.add(w[:i] + ch + w[i:])
        if ch in FOLDED_VOWELS:
            for v in FOLDED_VOWELS:
                if v != ch:
                    out.add(w[:i] + v + w[i+1:])
        if ch == 'j':
            out.add(w[:i] + w[i+1:])
    for i in range(L + 1):
        out.add(w[:i] + 'j' + w[i:])
    out.discard(w)
    return out


def lookup_fuzzy(ftok: str, forms_dict: dict) -> list[tuple[str, float, str, int]]:
    """Fuzzy lookup with operation-type ranking.
    Returns [(folded_dict_form, norm_dist, lemma, op_penalty)].
    Candidates sorted by (norm_dist, op_penalty, form)."""
    if len(ftok) < 3:
        return []
    all_forms = forms_dict["all_forms"]
    hits = set()
    for alt in ed1_neighbors(ftok, forms_dict["alphabet"]):
        if alt in all_forms:
            hits.add(alt)
    for v1 in plausible_variants(ftok):
        for v2 in plausible_variants(v1):
            if v2 != ftok and v2 in all_forms:
                hits.add(v2)

    candidates = []
    for df in hits:
        d = edit_distance(ftok, df)
        if d == 0:
            continue
        nd = d / max(len(ftok), len(df))
        penalty = op_penalty(ftok, df)
        lemma = forms_dict["folded_to_lemma"].get(df, ["?"])[0]
        candidates.append((df, nd, lemma, penalty))
    candidates.sort(key=lambda x: (x[1], x[3], x[0]))
    return candidates[:5]


# ── Analysis modules ──

def detect_proper_names(type_list: list) -> set:
    """Detect proper names: at least one medial occurrence with uppercase first char."""
    proper = set()
    for te in type_list:
        has_medial_upper = any(
            occ.form_with_case[0].isupper() and occ.position_type == "sentence_medial"
            for occ in te.occurrences
        )
        if has_medial_upper:
            proper.add(te.folded)
            te.is_proper_name = True
    return proper


PRUSSIAN_SUFFIXES = [
    "ans", "mai", "ins", "esse", "esmu", "imans", "amans",
    "essei", "ammans", "emmans", "ēisan", "ēimans",
    "īns", "ons", "uns", "asse", "ammai",
]
INTERJECTION_PATTERNS = [
    re.compile(r'^[aāeēiīoōuū]+h[aāeēiīoōuū]+$'),        # ahā, ehē, ohō
    re.compile(r'^[aāeēiīoōuū]{2,3}h?[aāeēiīoōuū]?$'),     # āh, ōhō
    re.compile(r'^[aāeēiīoōuū]{3,}$'),                      # āā, ūūū (3+ pure vowels)
    re.compile(r'^([aāeēiīoōuū])\1+i?$'),                   # āāā, ūūū
]


def detect_interjections(oov_entries: list) -> set:
    """Detect unmapped interjections in OOV list."""
    interj = set()
    for te in oov_entries:
        ft = te.folded
        for pat in INTERJECTION_PATTERNS:
            if pat.match(ft) and len(ft) >= 2 and len(ft) <= 6:
                interj.add(ft)
                te.interjection = True
                break
    return interj


def detect_loanword_endings(oov_entries: list) -> set:
    """Detect OOV words with known Prussian endings on unknown stems."""
    loan = set()
    for te in oov_entries:
        ft = te.folded
        if len(ft) < 6:
            continue
        for suffix in sorted(PRUSSIAN_SUFFIXES, key=len, reverse=True):
            if ft.endswith(suffix) and len(ft) - len(suffix) >= 4:
                stem = ft[:-len(suffix)]
                te.loanword_ending = suffix
                loan.add(ft)
                break
    return loan

SUSPICIOUS_SUFFIX_PAIRS = [
    ("tan", "tun", "Corpus -tan matched to -tun verb"),
    ("ten", "tun", "Corpus -ten matched to -tun verb"),
    ("tin", "tun", "Corpus -tin matched to -tun verb"),
    ("ton", "tun", "Corpus -ton matched to -tun verb"),
]


def detect_suspicious_matches(sig_candidates: list, forms_dict: dict) -> set:
    """Detect suspicious signature matches."""
    suspicious = set()
    lemma_info = forms_dict.get("lemma_info", {})

    for sc in sig_candidates:
        cf = fold(sc["type"].form)
        df = sc["best_folded"]
        lemma = sc["best_lemma"]
        fl = fold(lemma)
        info = lemma_info.get(fl, {})

        # Check suffix pairs
        for csuf, dsuf, reason in SUSPICIOUS_SUFFIX_PAIRS:
            if cf.endswith(csuf) and df.endswith(dsuf) and cf != df:
                if info.get("is_verb"):
                    sc["suspicious_pos"] = reason
                    suspicious.add((cf, df, reason))
                    break

        # Corpus form ends in -ais but dict lemma is a verb (imperative/adverbial confusion)
        if cf.endswith("ais") and info.get("is_verb"):
            if not sc.get("suspicious_pos"):
                sc["suspicious_pos"] = "Corpus -ais → verb (imperative?)"
                suspicious.add((cf, df, "Corpus -ais → verb"))

        # Corpus form ends in -an/-in-like suffix, dict lemma is -tun verb
        if (cf.endswith(("an", "in", "en")) and not df.endswith(("an", "in", "en"))
                and info.get("is_verb") and fl.endswith("tun")):
            if not sc.get("suspicious_pos"):
                sc["suspicious_pos"] = "Corpus nominal ending → -tun verb"
                suspicious.add((cf, df, "nominal → verb"))

    return suspicious


def detect_typos(sig_candidates: list, type_list: list) -> set:
    """Detect low-frequency probable typos (1-2 occ, ED=1, dict form well-attested)."""
    typos = set()
    type_freq = {te.folded: te.frequency for te in type_list}

    for sc in sig_candidates:
        te = sc["type"]
        if te.frequency <= 2:
            cf = te.folded
            best_df = sc["best_folded"]
            d = edit_distance(cf, best_df)
            if d <= 1:
                dict_freq = type_freq.get(best_df, 0)
                if dict_freq >= 5:
                    te.is_likely_typo = True
                    te.corpus_freq_of_match = dict_freq
                    typos.add(cf)
    return typos


# ── Error pattern classification ──

PATTERN_ORDER = [
    "Stem Vowel Deletion in Suffix",
    "Infinitive Ending (-tan → -tun)",
    "-ais Non-Verb",
    "Vowel Length (Macron)",
    "w-Prosthesis (w-Prefix)",
    "j-Palatalisation",
    "n-Insertion/Deletion",
    "Gemination (Single ↔ Double Consonant)",
    "Other Deletion",
    "Other Insertion",
    "Other Substitution",
    "Multi-Operation",
]


def _is_vlength(op: tuple) -> bool:
    return op[0] == 'sub' and VOWEL_MAP.get(op[1]) == op[2]


def classify_error_pattern(sc: dict, forms_dict: dict) -> str:
    """Map a signature candidate to a linguistic error pattern label."""
    td = sc["type"]
    cf = td.form.lower()
    df = sc["best_form"].lower()
    lemma = sc["best_lemma"]
    ops = sc["ops"]
    clen = len(cf)

    changes = [op for op in ops if op[0] != 'eq']
    if not changes:
        return "EXACT"

    # Separate macron/length changes from core (structural) changes
    vlength = [op for op in changes if _is_vlength(op)]
    core = [op for op in changes if op not in vlength]

    # Pure vowel-length: low-weight macron error
    if not core:
        return "Vowel Length (Macron)"

    n = len(core)

    # ── Specific multi-patterns first (ignore macrons) ──
    if cf.endswith(("tan", "ten", "tin", "ton")) and df.endswith("tun"):
        return "Infinitive Ending (-tan → -tun)"

    if cf.endswith("ais"):
        info = forms_dict.get("lemma_info", {}).get(fold(lemma), {})
        if info.get("is_verb"):
            return "-ais Non-Verb"

    # ── Single core operation patterns ──
    if n == 1:
        op, ca, cb, pa, pb = core[0]

        if op == 'del' and ca in VOWELS and clen - pa <= 4:
            return "Stem Vowel Deletion in Suffix"

        if op == 'ins' and cb == 'w' and pb == 0:
            return "w-Prosthesis (w-Prefix)"
        if op == 'del' and ca == 'w' and pa == 0:
            return "w-Prosthesis (w-Prefix)"

        if op == 'ins' and cb == 'j':
            return "j-Palatalisation"
        if op == 'del' and ca == 'j':
            return "j-Palatalisation"

        if op == 'ins' and cb == 'n':
            return "n-Insertion/Deletion"
        if op == 'del' and ca == 'n':
            return "n-Insertion/Deletion"

        if op == 'ins' and cb not in VOWELS:
            pcheck = len(df)
            if (pb > 0 and df[pb-1] == cb) or (pb + 1 < pcheck and df[pb+1] == cb):
                return "Gemination (Single ↔ Double Consonant)"
        if op == 'del' and ca not in VOWELS:
            if (pa > 0 and cf[pa-1] == ca) or (pa + 1 < clen and cf[pa+1] == ca):
                return "Gemination (Single ↔ Double Consonant)"

    # ── Single core op: generic ──
    if n == 1:
        op, ca, cb, pa, pb = core[0]
        if op == 'ins':
            return "Other Insertion"
        if op == 'del':
            return "Other Deletion"
        if op == 'sub':
            return "Other Substitution"

    return "Multi-Operation"


def build_pattern_groups(sig_candidates: list, forms_dict: dict) -> list:
    """Group signature candidates by linguistic error pattern."""
    groups = defaultdict(list)
    for sc in sig_candidates:
        pattern = classify_error_pattern(sc, forms_dict)
        sc["error_pattern"] = pattern
        groups[pattern].append(sc)

    pattern_list = []
    for pattern in PATTERN_ORDER:
        if pattern in groups:
            items = groups[pattern]
            items.sort(key=lambda x: -x["type"].frequency)
            pattern_list.append({
                "pattern": pattern,
                "total_frequency": sum(it["type"].frequency for it in items),
                "n_types": len(items),
                "items": items,
            })

    # Append any patterns not in PATTERN_ORDER
    for pattern, items in groups.items():
        if pattern not in PATTERN_ORDER:
            items.sort(key=lambda x: -x["type"].frequency)
            pattern_list.append({
                "pattern": pattern,
                "total_frequency": sum(it["type"].frequency for it in items),
                "n_types": len(items),
                "items": items,
            })

    for i, pg in enumerate(pattern_list, 1):
        pg["pattern_id"] = f"PAT-{i:03d}"

    return pattern_list


# ── Pipeline ──

PARTICIPLE_PAIRS = {
    ("wusis", "wuns"), ("wusi", "wuns"), ("wus", "wuns"),
    ("usis", "uns"), ("usi", "uns"),
    ("wunsi", "wuns"), ("unsi", "uns"),
}


def _is_grammatical_alt(sc: dict) -> bool:
    orig_corpus = sc["type"].form
    orig_dict = sc["best_form"]
    for suffix_c, suffix_d in PARTICIPLE_PAIRS:
        if orig_corpus.endswith(suffix_c) and orig_dict.endswith(suffix_d):
            return True
        if orig_dict.endswith(suffix_c) and orig_corpus.endswith(suffix_d):
            return True
    return False


def run_pipeline() -> dict:
    t_start = time.time()
    report = {}

    print("Loading dictionary full-forms...", file=sys.stderr)
    forms_dict = load_dict_forms()
    prec = check_precondition(forms_dict)
    report["precondition"] = {
        "all_pass": prec["all_pass"],
        "n_entries": prec["n_entries"],
        "n_lemmas": prec["n_lemmas"],
        "tested": [
            {"form": r["form"], "lemma": r["lemma"], "hit": r["hit"]}
            for r in prec["tested"]
        ],
    }
    if not prec["all_pass"]:
        print("  ⚠ PRECONDITION FAILED — some forms not found exactly!", file=sys.stderr)
    else:
        print(f"  ✅ Precondition PASS — {prec['n_entries']} folded forms, "
              f"{prec['n_lemmas']} lemmas", file=sys.stderr)

    print("Tokenizing corpus...", file=sys.stderr)
    type_list = tokenize_corpus()
    report["n_types_total"] = len(type_list)
    report["n_corpus_entries"] = len(json.loads(
        CORPUS_PATH.read_text(encoding="utf-8")))
    print(f"  {len(type_list)} types from corpus", file=sys.stderr)

    # ── 1. Proper name detection (runs first, before exact lookup) ──
    print("Detecting proper names...", file=sys.stderr)
    proper_names = detect_proper_names(type_list)
    report["n_proper_names"] = len(proper_names)
    print(f"  {len(proper_names)} proper name candidates", file=sys.stderr)

    # ── 2. FST-basierte Exact-Lookup (base.fst → lenient.fst fallback) ──
    print("FST exact lookup...", file=sys.stderr)
    fst_known = fst_lookup_types(type_list)

    known = []
    unknown = []
    for td in type_list:
        if td.folded in fst_known:
            lemma, analysis = fst_known[td.folded]
            td.dict_lemma = lemma
            known.append(td)
        else:
            unknown.append(td)
    report["n_exact_hits"] = len(known)
    print(f"  {len(known)} FST exact hits, {len(unknown)} unmatched", file=sys.stderr)

    # ── 3. Broken-Lemma-Routing ──
    print("Routing broken-lemma forms...", file=sys.stderr)
    broken_lemmas = forms_dict["broken_lemmas"]
    for b in broken_lemmas:
        b["corpus_forms"] = []
    broken_by_stem = sorted((b for b in broken_lemmas if b["stem"]),
                            key=lambda b: -len(b["stem"]))
    still_unknown = []
    for td in unknown:
        ftok = td.folded
        hit = None
        for b in broken_by_stem:
            st = b["stem"]
            if not (ftok.startswith(st) and len(ftok) - len(st) <= 6):
                continue
            rest = ftok[len(st):]
            if b["is_verb"] and rest and rest.endswith(("tun", "twei")):
                continue
            hit = b
            break
        if hit:
            hit["corpus_forms"].append(td)
        else:
            still_unknown.append(td)
    unknown = still_unknown
    report["n_broken_lemmas"] = len(broken_lemmas)
    report["n_broken_corpus_types"] = sum(
        len(b["corpus_forms"]) for b in broken_lemmas)
    print(f"  {len(broken_lemmas)} broken lemmas, "
          f"{report['n_broken_corpus_types']} corpus types routed", file=sys.stderr)

    # ── 4. Fuzzy lookup ──
    print("Fuzzy lookup...", file=sys.stderr)
    sig_candidates = []
    oov_candidates = []

    def _add_sig_candidate(td, best_df, best_nd, best_lemma, forms_dict):
        orig_corpus = unicodedata.normalize("NFC", td.form).lower()
        orig_dict = unicodedata.normalize("NFC",
            forms_dict["folded_to_orig"].get(best_df, best_df)).lower()
        ops = edit_ops(orig_corpus, orig_dict)
        sig = derive_signature(ops, orig_corpus, orig_dict)
        f_corpus = fold(orig_corpus)
        f_dict = fold(orig_dict)
        d = edit_distance(f_corpus, f_dict)
        nd = d / max(len(f_corpus), len(f_dict), 1)
        return {
            "type": td,
            "best_folded": best_df,
            "best_form": orig_dict,
            "best_lemma": best_lemma,
            "norm_dist": nd,
            "ops": ops,
            "signature": sig,
            "edit_distance": d,
            "suspicious_pos": "",
        }

    dict_form_list = sorted(forms_dict["all_forms"])

    def _nearest_neighbor(ftok: str) -> tuple:
        if not HAVE_RAPIDFUZZ:
            return None, None, None, None
        m = _rf_process.extractOne(ftok, dict_form_list,
                                   scorer=_Lev.normalized_distance)
        if not m:
            return None, None, None, None
        nn, nd, _idx = m
        return (nn,
                forms_dict["folded_to_orig"].get(nn, nn),
                forms_dict["folded_to_lemma"].get(nn, ["?"])[0],
                nd)

    for td in unknown:
        ftok = td.folded
        if len(ftok) <= 2:
            oov_candidates.append({"type": td, "nearest_folded": None,
                                    "nearest_form": None, "nearest_lemma": None,
                                    "nearest_dist": None})
            continue

        cands = lookup_fuzzy(ftok, forms_dict)
        if cands:
            best_df, best_nd, best_lemma, best_penalty = cands[0]
            sc = _add_sig_candidate(td, best_df, best_nd, best_lemma, forms_dict)
            sig_candidates.append(sc)
        else:
            nn_folded, nn_form, nn_lemma, nn_dist = _nearest_neighbor(ftok)
            oov_candidates.append({
                "type": td,
                "nearest_folded": nn_folded,
                "nearest_form": nn_form,
                "nearest_lemma": nn_lemma,
                "nearest_dist": nn_dist,
            })

    report["n_signature_raw"] = len(sig_candidates)
    report["n_oov"] = len(oov_candidates)
    print(f"  {len(sig_candidates)} signature candidates, {len(oov_candidates)} OOV",
          file=sys.stderr)

    # ── 5. Filter known grammatical alternations ──
    print("Filtering known grammatical alternations...", file=sys.stderr)
    sig_candidates = [sc for sc in sig_candidates if not _is_grammatical_alt(sc)]
    n_filtered = report["n_signature_raw"] - len(sig_candidates)
    report["n_signature"] = len(sig_candidates)
    if n_filtered:
        print(f"  Filtered {n_filtered} grammatical alternations (participles)",
              file=sys.stderr)

    # ── 6. Analysis passes ──
    print("Running analysis passes...", file=sys.stderr)

    # 6a. Typo detection
    oov_tds = [oc["type"] for oc in oov_candidates]
    n_typos = detect_typos(sig_candidates, type_list)
    print(f"  Typo candidates: {len(n_typos)}", file=sys.stderr)
    report["n_typos"] = len(n_typos)

    # 6b. Suspicious POS mismatch
    n_suspicious = detect_suspicious_matches(sig_candidates, forms_dict)
    print(f"  Suspicious POS matches: {len(n_suspicious)}", file=sys.stderr)
    report["n_suspicious"] = len(n_suspicious)

    # 6c. Interjection detection
    n_interjections = detect_interjections(oov_tds)
    print(f"  Interjection candidates: {len(n_interjections)}", file=sys.stderr)
    report["n_interjections"] = len(n_interjections)

    # 6d. Loanword ending detection
    n_loanwords = detect_loanword_endings(oov_tds)
    print(f"  Loanword-ending candidates: {len(n_loanwords)}", file=sys.stderr)
    report["n_loanwords"] = len(n_loanwords)

    # ── 7. Cluster by signature (mechanical) ──
    print("Clustering by signature...", file=sys.stderr)
    sig_clusters = defaultdict(list)
    for sc in sig_candidates:
        sig_clusters[sc["signature"]].append(sc)

    cluster_list = []
    for sig, items in sig_clusters.items():
        items.sort(key=lambda x: -x["type"].frequency)
        cluster_list.append({
            "signature": sig,
            "total_frequency": sum(it["type"].frequency for it in items),
            "n_types": len(items),
            "items": items,
        })
    cluster_list.sort(key=lambda x: -x["total_frequency"])

    for i, cl in enumerate(cluster_list, 1):
        cl["sig_id"] = f"SIG-{i:03d}"

    report["n_clusters"] = len(cluster_list)

    # ── 8. Build pattern groups ──
    print("Building error pattern groups...", file=sys.stderr)
    pattern_groups = build_pattern_groups(sig_candidates, forms_dict)
    report["n_pattern_groups"] = len(pattern_groups)
    print(f"  {len(pattern_groups)} pattern groups", file=sys.stderr)

    oov_list = sorted(oov_candidates, key=lambda x: -x["type"].frequency)
    report["oov_list"] = oov_list
    report["cluster_list"] = cluster_list
    report["pattern_groups"] = pattern_groups
    report["broken_lemmas"] = broken_lemmas
    report["proper_names"] = proper_names
    report["type_list"] = type_list
    report["match_policy"] = MATCH_POLICY
    report["elapsed"] = time.time() - t_start
    report["forms_dict"] = forms_dict

    return report


# ── Error group builder ──

def build_error_groups(report: dict) -> list:
    """Extract error groups from pipeline results."""
    groups = []

    # Group: Proper names in OOV
    proper_oov = [oc for oc in report["oov_list"] if oc["type"].is_proper_name]
    if proper_oov:
        groups.append({
            "group": "Potential proper names (OOV)",
            "type": "OOV + capitalization pattern",
            "count": len(proper_oov),
            "examples": ", ".join(oc["type"].form for oc in proper_oov[:10]),
            "explanation": "Uppercase in sentence-medial position suggests proper noun not in dictionary",
        })

    # Group: Proper names in signatures
    proper_sig = []
    for cl in report["cluster_list"]:
        for it in cl["items"]:
            if it["type"].is_proper_name:
                proper_sig.append(it)
    if proper_sig:
        groups.append({
            "group": "Potential proper names (signature matches)",
            "type": "Sig-match + capitalization",
            "count": len(proper_sig),
            "examples": ", ".join(p["type"].form for p in proper_sig[:10]),
            "explanation": "Capitalized forms matched to dictionary entries via edit ops; may be proper names",
        })

    # Group: Typos
    typo_items = []
    for cl in report["cluster_list"]:
        for it in cl["items"]:
            if it["type"].is_likely_typo:
                typo_items.append(it)
    if typo_items:
        groups.append({
            "group": "Probable typos (1-2 occ, dict form well-attested)",
            "type": "Low-freq typo",
            "count": len(typo_items),
            "examples": ", ".join(
                f"{p['type'].form}→{p['best_form']}" for p in typo_items[:10]),
            "explanation": "Low-frequency forms within ED=1 of high-frequency dictionary forms",
        })

    # Group: Interjections
    interj = [oc for oc in report["oov_list"] if oc["type"].interjection]
    if interj:
        groups.append({
            "group": "Unmapped interjections",
            "type": "OOV + vowel pattern",
            "count": len(interj),
            "examples": ", ".join(oc["type"].form for oc in interj[:10]),
            "explanation": "Short vowel-heavy forms not in dictionary; likely interjections/exclamations",
        })

    # Group: Loanword endings
    loan = [oc for oc in report["oov_list"] if oc["type"].loanword_ending]
    if loan:
        groups.append({
            "group": "OOV words with known Prussian endings",
            "type": "Loanword candidate",
            "count": len(loan),
            "examples": ", ".join(
                f"{oc['type'].form} (+{oc['type'].loanword_ending})"
                for oc in loan[:10]),
            "explanation": "Words ending in Prussian declension suffixes but stem not in dictionary",
        })

    # Group: Suspicious POS matches (suffix-based)
    susp_items = []
    for cl in report["cluster_list"]:
        for it in cl["items"]:
            if it.get("suspicious_pos", ""):
                susp_items.append(it)
    if susp_items:
        susp_by_reason = defaultdict(list)
        for it in susp_items:
            reason = it.get("suspicious_pos", "Unknown")
            susp_by_reason[reason].append(it)
        for reason, items in susp_by_reason.items():
            groups.append({
                "group": f"Suspicious POS: {reason}",
                "type": "Suffix mismatch",
                "count": len(items),
                "examples": ", ".join(
                    f"{it['type'].form}→{it['best_form']}" for it in items[:10]),
                "explanation": reason,
            })

    return groups


# ── XLSX styling ──

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(size=10)
LINK_FONT = Font(size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
GROUP_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TYPO_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
SUSP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PROPER_FONT = Font(size=10, color="2F5496")


def highlight_form_in_context(context: str, form: str) -> CellRichText:
    """Return CellRichText with the *form* substring in red+bold."""
    idx = context.lower().find(form.lower())
    if idx < 0:
        return CellRichText(TextBlock(InlineFont(), context))

    before = context[:idx]
    match_str = context[idx:idx + len(form)]
    after = context[idx + len(form):]

    parts = []
    if before:
        parts.append(TextBlock(InlineFont(), before))
    parts.append(TextBlock(InlineFont(color="FF0000", b=True), match_str))
    if after:
        parts.append(TextBlock(InlineFont(), after))
    return CellRichText(*parts)


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        cell.border = THIN_BORDER


def _style_body(ws, nrows: int, ncols: int):
    for row in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _col_letter(i: int) -> str:
    """Column letter for 1-based index (supports >26)."""
    result = ""
    while i > 0:
        i, rem = divmod(i - 1, 26)
        result = chr(65 + rem) + result
    return result


def write_signatures_xlsx(pattern_groups: list, path: Path):
    """One row per variant; grouped by linguistic error pattern.
    Pattern groups get a header row. Corpus form highlighted red in context."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Signatures"
    headers = [
        "Error Pattern", "Dict Form", "Lemma",
        "Freq", "Example Sentence", "Youtubebe Links",
        "Mech. Signature", "Proper Name", "Likely Typo",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    PATTERN_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496",
                                      fill_type="solid")
    PATTERN_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

    ri = 2
    for pg in pattern_groups:
        # Pattern group header row
        ws.cell(row=ri, column=1,
                value=f"▸ {pg['pattern']} ({pg['n_types']} types, Σ freq {pg['total_frequency']})")
        for col in range(2, len(headers) + 1):
            ws.cell(row=ri, column=col, value="")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=col)
            cell.fill = PATTERN_HEADER_FILL
            cell.font = PATTERN_HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(headers))
        ri += 1

        for it in pg["items"]:
            td = it["type"]
            ctx = td.example_context()
            links = td.top_links(3)
            ctx_rich = highlight_form_in_context(ctx, td.form)

            ws.cell(row=ri, column=1, value=pg["pattern"])
            ws.cell(row=ri, column=2, value=it["best_form"])
            ws.cell(row=ri, column=3, value=it["best_lemma"])
            ws.cell(row=ri, column=4, value=td.frequency)
            ws.cell(row=ri, column=5, value=ctx_rich)
            ws.cell(row=ri, column=6, value=links)
            ws.cell(row=ri, column=7, value=it["signature"])
            ws.cell(row=ri, column=8, value="✓" if td.is_proper_name else "")
            ws.cell(row=ri, column=9,
                    value=f"✓ (cf. {td.corpus_freq_of_match})" if td.is_likely_typo else "")

            if td.is_likely_typo:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ri, column=col).fill = TYPO_FILL
            elif it.get("suspicious_pos", ""):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ri, column=col).fill = SUSP_FILL
            elif td.is_proper_name:
                ws.cell(row=ri, column=5).font = PROPER_FONT

            ri += 1

    nrows = ri - 2
    _style_body(ws, nrows, len(headers))

    last_col = _col_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{nrows + 1}"
    ws.freeze_panes = "A2"

    widths = [30, 18, 16, 6, 55, 45, 34, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  📄 {path}", file=sys.stderr)


def write_oov_xlsx(oov_list: list, path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OOV"
    headers = [
        "Corpus Form", "Frequency",
        "Context", "Source Links",
        "Nearest Dict Neighbor", "Nearest Dist (norm.)",
        "Proper Name", "Interjection", "Loanword Ending",
        "Glabbis: Lemma", "Glabbis: Comment",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    for ri, oc in enumerate(oov_list, 2):
        td = oc["type"]
        ws.cell(row=ri, column=1, value=td.form)
        ws.cell(row=ri, column=2, value=td.frequency)
        ws.cell(row=ri, column=3, value=td.example_context())
        ws.cell(row=ri, column=4, value=td.top_links(3))
        if oc.get("nearest_form") and oc.get("nearest_dist") is not None:
            ws.cell(row=ri, column=5,
                    value=f"{oc['nearest_form']} (Lemma: {oc['nearest_lemma']})")
            ws.cell(row=ri, column=6, value=round(oc["nearest_dist"], 3))
        else:
            ws.cell(row=ri, column=5, value="")
            ws.cell(row=ri, column=6, value="")
        ws.cell(row=ri, column=7, value="✓" if td.is_proper_name else "")
        ws.cell(row=ri, column=8, value="✓" if td.interjection else "")
        ws.cell(row=ri, column=9, value=td.loanword_ending or "")
        ws.cell(row=ri, column=10, value="")
        ws.cell(row=ri, column=11, value="")

        if td.is_proper_name:
            ws.cell(row=ri, column=1).font = PROPER_FONT
        if td.interjection:
            ws.cell(row=ri, column=1).fill = SUSP_FILL

    _style_body(ws, len(oov_list), len(headers))
    last_col = _col_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(oov_list) + 1}"
    ws.freeze_panes = "A2"

    widths = [22, 10, 50, 45, 35, 14, 12, 12, 16, 20, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  📄 {path}", file=sys.stderr)


def write_error_groups_xlsx(groups: list, path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Error Groups"
    headers = ["Error Group", "Type", "#Entries", "Examples", "Explanation"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    for ri, g in enumerate(groups, 2):
        ws.cell(row=ri, column=1, value=g["group"])
        ws.cell(row=ri, column=2, value=g["type"])
        ws.cell(row=ri, column=3, value=g["count"])
        ws.cell(row=ri, column=4, value=g["examples"])
        ws.cell(row=ri, column=5, value=g["explanation"])

    _style_body(ws, len(groups), len(headers))
    ws.auto_filter.ref = f"A1:E{len(groups) + 1}"
    ws.freeze_panes = "A2"

    widths = [42, 24, 10, 80, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  📄 {path}", file=sys.stderr)


def write_broken_xlsx(broken_lemmas: list, path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Broken Lemmas"
    headers = [
        "Lemma", "Paradigm", "#Cells", "Corpus Forms (Freq)",
        "#Tokens in Corpus", "Comment",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    rows = sorted(broken_lemmas,
                  key=lambda b: (-sum(td.frequency for td in b["corpus_forms"]),
                                 b["word"]))
    for ri, b in enumerate(rows, 2):
        n_tok = sum(td.frequency for td in b["corpus_forms"])
        forms = ", ".join(
            f"{td.form} ({td.frequency})"
            for td in sorted(b["corpus_forms"], key=lambda t: -t.frequency))
        ws.cell(row=ri, column=1, value=b["word"])
        ws.cell(row=ri, column=2, value=b["paradigm"])
        ws.cell(row=ri, column=3, value=b["n_cells"])
        ws.cell(row=ri, column=4, value=forms)
        ws.cell(row=ri, column=5, value=n_tok)
        ws.cell(row=ri, column=6, value="")

    _style_body(ws, len(rows), len(headers))
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"
    ws.freeze_panes = "A2"

    widths = [24, 12, 10, 60, 16, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  📄 {path}", file=sys.stderr)


# ── Main ──

def main():
    report = run_pipeline()

    # Build error groups
    print("\nBuilding error groups...", file=sys.stderr)
    error_groups = build_error_groups(report)

    # Output XLSX
    print("Writing XLSX reports...", file=sys.stderr)
    write_signatures_xlsx(report["pattern_groups"], OUT_SIG)
    write_oov_xlsx(report["oov_list"], OUT_OOV)
    write_broken_xlsx(report["broken_lemmas"], OUT_BROKEN)
    write_error_groups_xlsx(error_groups, OUT_ERR_GROUPS)

    r = report
    print(f"\n{'='*60}", file=sys.stderr)
    print("RUN REPORT", file=sys.stderr)
    print(f"{'='*60}", file=sys.stderr)
    print(f"  Precondition:      {'PASS' if r['precondition']['all_pass'] else 'FAIL'}",
          file=sys.stderr)
    print(f"  Dict folded forms: {r['precondition']['n_entries']}", file=sys.stderr)
    print(f"  Dict lemmas:       {r['precondition']['n_lemmas']}", file=sys.stderr)
    print(f"  Corpus entries:    {r['n_corpus_entries']}", file=sys.stderr)
    print(f"  Types total:       {r['n_types_total']}", file=sys.stderr)
    print(f"  Proper names:      {r['n_proper_names']}", file=sys.stderr)
    print(f"  Exact hits:        {r['n_exact_hits']}", file=sys.stderr)
    print(f"  Sig-instances:     {r['n_signature']}", file=sys.stderr)
    print(f"  Typos detected:    {r['n_typos']}", file=sys.stderr)
    print(f"  Suspicious POS:    {r['n_suspicious']}", file=sys.stderr)
    print(f"  Interjections:     {r['n_interjections']}", file=sys.stderr)
    print(f"  Loanword endings:  {r['n_loanwords']}", file=sys.stderr)
    print(f"  OOV:               {r['n_oov']}", file=sys.stderr)
    print(f"  Broken lemmas:     {r['n_broken_lemmas']} "
          f"({r['n_broken_corpus_types']} corpus types)", file=sys.stderr)
    print(f"  Pattern groups:    {r['n_pattern_groups']}", file=sys.stderr)
    print(f"  Error groups:      {len(error_groups)}", file=sys.stderr)
    print(f"  Match-Politik:     {r['match_policy']}", file=sys.stderr)
    print(f"  Elapsed:           {r['elapsed']:.1f}s", file=sys.stderr)

    print(f"\n  Pattern groups:", file=sys.stderr)
    print(f"  {'Pattern':<38} {'Freq':>6} {'#Type':>6}", file=sys.stderr)
    print(f"  {'-'*38} {'-'*6} {'-'*6}", file=sys.stderr)
    for pg in r["pattern_groups"]:
        print(f"  {pg['pattern'][:36]:<38} {pg['total_frequency']:>6} {pg['n_types']:>6}",
              file=sys.stderr)

    summary = {
        "precondition_pass": r["precondition"]["all_pass"],
        "n_dict_folded_forms": r["precondition"]["n_entries"],
        "n_dict_lemmas": r["precondition"]["n_lemmas"],
        "n_types_total": r["n_types_total"],
        "n_exact_hits": r["n_exact_hits"],
        "n_signature_instances": r["n_signature"],
        "n_oov": r["n_oov"],
        "n_broken_lemmas": r["n_broken_lemmas"],
        "n_broken_corpus_types": r["n_broken_corpus_types"],
        "n_proper_names": r["n_proper_names"],
        "n_typos": r["n_typos"],
        "n_suspicious": r["n_suspicious"],
        "n_interjections": r["n_interjections"],
        "n_loanwords": r["n_loanwords"],
        "n_pattern_groups": r["n_pattern_groups"],
        "n_error_groups": len(error_groups),
        "match_policy": r["match_policy"],
        "elapsed_seconds": round(r["elapsed"], 1),
        "output_signatures": str(OUT_SIG),
        "output_oov": str(OUT_OOV),
        "output_broken": str(OUT_BROKEN),
        "output_error_groups": str(OUT_ERR_GROUPS),
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
