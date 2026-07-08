#!/usr/bin/env python3
"""Gemeinsame Bausteine für Korpus↔Wörterbuch-Abgleich (Tokenizer,
FST-Batch-Lookup, Dictionary-Loader, xlsx-Styling-Helfer).

Wird von delta_review.py und error_classification.py importiert, damit
Tokenisierung/Lookup nicht zweimal gepflegt werden müssen.
"""

import json
import re
import subprocess
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

try:
    from rapidfuzz.distance import Levenshtein as _Lev
    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

REPO = Path(__file__).resolve().parents[2]
TWANKSTA_PATH = REPO / "../prussian-corpus/parsed/twanksta_entries.json"
CORPUS_PATH = REPO / "../prussian-corpus/parsed/youtube_corpus_sentences.json"

FST_DIR = Path(__file__).resolve().parents[1]
BASE_FST = FST_DIR / "build/base.fst"
LENIENT_FST = FST_DIR / "build/lenient.fst"
RULES_DIR = FST_DIR / "build/rules"

# Ortho-Layer-Einzelregeln (s. fst/norm.regex) mit menschenlesbarem Label.
# Reihenfolge ist die Prüfreihenfolge, falls mehrere gleichzeitig passen.
ORTHO_RULES = [
    ("degem_all", "Degemination"),
    ("w_prothese", "w-Prosthesis"),
    ("samland_i", "Samland i-Epenthesis"),
    ("i_synkope", "i-Syncope (Nom.Sg. i-stems)"),
    ("diph_ai", "Diphthong Simplification aī→ai"),
    ("macron_loss", "Macron Loss"),
    ("inf_tun_twei", "Infinitive -twei/-tun"),
]

SKIP_VIDEOS = {"qLwBCWtMuH8"}


@dataclass
class SourceRef:
    video_id: str
    title: str
    start: str   # "00:02:19,300"
    end: str
    translation: str = ""

    @property
    def t_seconds(self) -> int:
        try:
            clean = self.start.replace(",", ".")
            h, m, s = clean.split(":")
            return int(h) * 3600 + int(m) * 60 + int(float(s))
        except (ValueError, AttributeError):
            return 0

    @property
    def youtube_url(self) -> str:
        ts = self.t_seconds
        return f"https://youtu.be/{self.video_id}?t={ts}" if ts else f"https://youtu.be/{self.video_id}"


@dataclass
class Occurrence:
    context: str       # full sentence text
    position: int      # 0-based index of this token in sentence
    position_type: str  # "sentence_initial", "sentence_medial", "only_token"
    form_with_case: str  # original casing as it appeared here
    sentence_frequency: int
    sources: list = field(default_factory=list)  # list of SourceRef
    translations: list = field(default_factory=list)  # [(text, count), ...]


@dataclass
class TypeEntry:
    form: str            # canonical form (first encounter, preserves case)
    folded: str
    frequency: int       # total across all sentences × positions
    occurrences: list = field(default_factory=list)  # list of Occurrence
    dict_lemma: str = ""
    is_proper_name: bool = False

    @property
    def source_links(self) -> str:
        urls = []
        seen = set()
        for occ in self.occurrences:
            for sr in occ.sources:
                key = (sr.video_id, sr.t_seconds)
                if key not in seen:
                    seen.add(key)
                    urls.append(sr.youtube_url)
        return "\n".join(urls)

    def pick_best_occurrence(self):
        """Return the occurrence with most sources (best linked example)."""
        if not self.occurrences:
            return None
        return max(self.occurrences,
                   key=lambda o: len(o.sources) if o.sources else 0)

    def top_links(self, n: int = 3) -> str:
        """First N unique youtube links across occurrences."""
        urls = []
        seen = set()
        for occ in self.occurrences:
            for sr in occ.sources:
                key = (sr.video_id, sr.t_seconds)
                if key not in seen:
                    seen.add(key)
                    urls.append(sr.youtube_url)
                    if len(urls) >= n:
                        return "\n".join(urls)
        return "\n".join(urls)

    def example_context(self) -> str:
        """One example sentence for display (prefer one with youtube sources)."""
        best = self.pick_best_occurrence()
        if best:
            return best.context
        return ""

    def example_snippet(self, window: int = 40) -> str:
        """Short excerpt centered on the matched form, not the full sentence
        — long corpus sentences otherwise blow up the review table."""
        ctx = self.example_context()
        if not ctx:
            return ""
        idx = ctx.lower().find(self.form.lower())
        if idx < 0:
            snippet = ctx
        else:
            start = max(0, idx - window)
            end = min(len(ctx), idx + len(self.form) + window)
            snippet = ctx[start:end]
            if start > 0:
                snippet = "…" + snippet
            if end < len(ctx):
                snippet = snippet + "…"
        return " ".join(snippet.split())

    def translation_hint(self) -> str:
        """Single most frequent translation, truncated — one hint is enough
        for manual review, a wall of duplicate translations is not."""
        counts = defaultdict(int)
        for occ in self.occurrences:
            for text, count in occ.translations:
                if text:
                    counts[text] += count
        if not counts:
            return ""
        best_text = max(counts.items(), key=lambda kv: kv[1])[0]
        best_text = " ".join(best_text.split())
        if len(best_text) > 80:
            best_text = best_text[:79] + "…"
        return best_text

    def translation_hints(self, n: int = 5) -> list[str]:
        """Aggregated, frequency-sorted translation strings across all
        occurrences (dedup by text) — used internally for signal detection,
        not for display (see translation_hint())."""
        counts = defaultdict(int)
        for occ in self.occurrences:
            for text, count in occ.translations:
                if text:
                    counts[text] += count
        ranked = sorted(counts.items(), key=lambda kv: -kv[1])
        return [text for text, _ in ranked[:n]]


def fold(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return unicodedata.normalize("NFC", s.lower())


def extract_prussian_tokens_with_pos(text: str) -> list[tuple[str, int, int]]:
    """Return (token, position, first_alpha_position) tuples.
    position_type is determined later by comparing position to
    first_alpha_position."""
    if re.match(r'^[A-Z]{2,5}:', text):
        if not text.startswith('PR:'):
            return []
        text = text[3:].lstrip()
    text = text.split("//")[0]
    text = re.sub(r'\[[^\]]*\]', '', text)
    text = re.sub(r'\([^)]*\)', '', text)
    tokens = []
    pos = 0
    first_alpha = None
    for tok in text.split():
        tok = tok.lstrip('=/')
        tok = tok.strip('.,!?;:()[]{}«»""\' \t')
        if tok and len(tok) >= 2 and tok.isalpha():
            if first_alpha is None:
                first_alpha = pos
            tokens.append((tok, pos, first_alpha if first_alpha is not None else 0))
        pos += 1
    return tokens


def edit_distance(a: str, b: str) -> int:
    if HAVE_RAPIDFUZZ:
        return _Lev.distance(a, b)
    m, n = len(a), len(b)
    if m < n:
        a, b = b, a
        m, n = n, m
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[0]
        dp[0] = i
        for j in range(1, n + 1):
            temp = dp[j]
            dp[j] = min(
                prev + (0 if a[i-1] == b[j-1] else 1),
                dp[j] + 1,
                dp[j-1] + 1,
            )
            prev = temp
    return dp[n]


# ── FST-based lookup ──

def fst_lookup_batch(forms: list[str], fst_path: Path) -> dict[str, list[tuple[str, str, list[str]]]]:
    """Batch lookup through hfst-flookup.
    Returns {form: [(lemma, analysis_str, tags), ...]}"""
    if not forms:
        return {}
    proc = subprocess.run(
        ["hfst-flookup", "-q", str(fst_path)],
        input="\n".join(forms) + "\n",
        capture_output=True, text=True, check=True,
    )
    results = defaultdict(list)
    for line in proc.stdout.splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        form, analysis = parts[0], parts[1]
        if analysis.endswith("+?"):
            continue
        segs = analysis.split("+")
        lemma, tags = segs[0], segs[1:]
        if not tags:
            continue
        results[form].append((lemma, analysis, tags))
    return dict(results)


def fst_lookup_types(type_list: list, base_fst: Path = BASE_FST,
                      lenient_fst: Path = LENIENT_FST) -> dict[str, tuple[str, str]]:
    """FST-based exact lookup: base.fst → lenient.fst fallback.
    Returns {folded_form: (lemma, analysis_str)} for known types."""
    forms_set = set()
    form_to_folded = {}  # original string → folded

    for te in type_list:
        forms_set.add(te.form)
        form_to_folded[te.form] = te.folded
        lo = te.form.lower()
        if lo != te.form:
            forms_set.add(lo)
            form_to_folded.setdefault(lo, te.folded)
        for occ in te.occurrences:
            fwc = occ.form_with_case
            if fwc not in forms_set:
                forms_set.add(fwc)
                form_to_folded.setdefault(fwc, te.folded)

    all_forms = sorted(forms_set)

    print(f"  FST lookup: {len(all_forms)} unique strings on base.fst...", file=sys.stderr)
    known = fst_lookup_batch(all_forms, base_fst)
    unknown = [f for f in all_forms if f not in known]

    if unknown and lenient_fst is not None:
        print(f"  FST fallback: {len(unknown)} onto lenient.fst...", file=sys.stderr)
        lenient = fst_lookup_batch(unknown, lenient_fst)
        for f, analyses in lenient.items():
            known[f] = analyses

    result = {}
    for orig_form, analyses in known.items():
        folded = form_to_folded.get(orig_form)
        if folded and folded not in result:
            lemma, analysis, tags = analyses[0]
            result[folded] = (lemma, analysis)

    return result


def fst_lookup_exact_only(forms: list[str], base_fst: Path = BASE_FST) -> dict[str, tuple[str, str]]:
    """Exact base.fst lookup only (no lenient fallback), keyed by the exact
    input string (not folded) — used to test individual ortho-rule
    transducers against corpus-attested spellings."""
    known = fst_lookup_batch(forms, base_fst)
    result = {}
    for form, analyses in known.items():
        lemma, analysis, tags = analyses[0]
        result[form] = (lemma, analysis)
    return result


# ── Dictionary loading ──

def load_dict_forms() -> dict:
    raw = json.loads(TWANKSTA_PATH.read_text(encoding="utf-8"))
    all_forms = set()
    forms_by_len = defaultdict(set)
    folded_to_lemma = defaultdict(list)
    folded_to_orig = {}
    lemma_by_form = {}
    lemma_info = {}  # folded_lemma → {"is_verb": bool, "paradigm": str, "gender": str}
    lemma_gloss = {}  # folded_lemma → list of english glosses

    def _clean(f: str) -> str:
        f = f.strip().strip('!?.,;: ')
        if not f or " " in f or "/" in f:
            return ""
        if not all(c.isalpha() or c == '-' for c in f):
            return ""
        return f

    def _add(f: str, lemma: str):
        f = _clean(f)
        if not f:
            return
        ff = fold(f)
        all_forms.add(ff)
        forms_by_len[len(ff)].add(ff)
        folded_to_lemma[ff].append(lemma)
        folded_to_orig.setdefault(ff, f)
        lemma_by_form.setdefault(f, lemma)

    def _broken_stem(word: str):
        fw = fold(word)
        if " " in fw:
            return None
        is_verb = False
        for suf in ("twei", "tun"):
            if fw.endswith(suf):
                fw = fw[:-len(suf)]
                is_verb = True
                break
        else:
            fw = fw.rstrip('aeiou')
        if len(fw) < 4:
            return None
        return (fw, is_verb)

    broken_lemmas = []

    for e in raw:
        word = e.get("word", "")
        if not word:
            continue
        _add(word, word)

        paradigm = e.get("paradigm", "")
        gender = e.get("gender", "")
        has_ind = bool(e.get("forms", {}).get("indicative"))
        has_opt = bool(e.get("forms", {}).get("optative"))
        has_imp = bool(e.get("forms", {}).get("imperative"))
        is_verb = has_ind or has_opt or has_imp

        fl = fold(word)
        lemma_info[fl] = {
            "is_verb": is_verb,
            "paradigm": paradigm,
            "gender": gender,
        }
        lemma_gloss[fl] = e.get("translations", {}).get("engl", [])

        decl_cells = []
        for g in e.get("forms", {}).get("declension", []):
            for c in g.get("cases", []):
                for num in ("singular", "plural"):
                    if c.get(num):
                        decl_cells.append(c[num])
        if len(decl_cells) >= 4 and set(decl_cells) == {word}:
            stem_info = _broken_stem(word)
            broken_lemmas.append({
                "word": word,
                "paradigm": paradigm,
                "n_cells": len(decl_cells),
                "stem": stem_info[0] if stem_info else None,
                "is_verb": stem_info[1] if stem_info else False,
            })
        else:
            for f in decl_cells:
                _add(f, word)

        for p in e.get("forms", {}).get("participles", []):
            _add(p.get("form", ""), word)

        for mood in ("indicative", "optative", "imperative", "subjunctive"):
            val = e.get("forms", {}).get(mood)
            if isinstance(val, list):
                for tense_entry in val:
                    if isinstance(tense_entry, dict):
                        for entry in tense_entry.get("forms", []):
                            _add(entry.get("form", ""), word)
                        _add(tense_entry.get("form", ""), word)
            elif isinstance(val, str):
                _add(val, word)

    folded_to_orig_form = {}
    for ff in all_forms:
        orig = folded_to_orig.get(ff)
        if orig:
            folded_to_orig_form[ff] = orig

    forms_by_len = {k: sorted(v) for k, v in forms_by_len.items()}
    n_lemmas = len(raw)

    return {
        "all_forms": all_forms,
        "alphabet": sorted({c for f in all_forms for c in f if c.isalpha()}),
        "forms_by_len": forms_by_len,
        "folded_to_lemma": dict(folded_to_lemma),
        "folded_to_orig": folded_to_orig,
        "folded_to_orig_form": folded_to_orig_form,
        "lemma_by_form": lemma_by_form,
        "n_lemmas": n_lemmas,
        "broken_lemmas": broken_lemmas,
        "lemma_info": lemma_info,
        "lemma_gloss": lemma_gloss,
    }


def check_precondition(forms_dict: dict) -> dict:
    lemma_samples = list(forms_dict["lemma_by_form"].items())
    inflected = [(f, lm) for f, lm in lemma_samples if f != lm and len(f) >= 3]
    test = inflected[:10]
    results = []
    for form, lemma in test:
        folded = fold(form)
        hit = folded in forms_dict["folded_to_lemma"]
        results.append({"form": form, "lemma": lemma, "folded": folded, "hit": hit})
    all_hit = all(r["hit"] for r in results)
    return {"tested": results, "all_pass": all_hit,
            "n_entries": len(forms_dict["all_forms"]),
            "n_original_forms": len(forms_dict["lemma_by_form"]),
            "n_lemmas": forms_dict["n_lemmas"]}


# ── Tokenize with occurrence tracking ──

def tokenize_corpus(corpus_path: Path = CORPUS_PATH,
                     skip_videos: set = SKIP_VIDEOS) -> list:
    raw = json.loads(corpus_path.read_text(encoding="utf-8"))
    type_data = {}  # folded → TypeEntry

    for e in raw:
        sources = e.get("sources", [])
        if sources and all(s.get("video_id") in skip_videos for s in sources):
            continue

        text = e.get("text", "")
        sent_freq = e.get("frequency", 1)
        translations = [(t.get("text", ""), t.get("count", 1))
                        for t in e.get("translations", []) if t.get("text")]

        source_refs = [
            SourceRef(
                video_id=s.get("video_id", ""),
                title=s.get("title", ""),
                start=s.get("start", "00:00:00,000"),
                end=s.get("end", "00:00:00,000"),
                translation=s.get("translation") or "",
            )
            for s in sources
        ]

        tokens = extract_prussian_tokens_with_pos(text)
        for tok, pos, first_alpha in tokens:
            ftok = fold(tok)
            if ftok not in type_data:
                type_data[ftok] = TypeEntry(
                    form=tok, folded=ftok, frequency=0, occurrences=[])

            te = type_data[ftok]
            te.frequency += sent_freq

            pos_type = "sentence_initial" if pos == first_alpha else "sentence_medial"

            te.occurrences.append(Occurrence(
                context=text,
                position=pos,
                position_type=pos_type,
                form_with_case=tok,
                sentence_frequency=sent_freq,
                sources=source_refs,
                translations=translations,
            ))

    return sorted(type_data.values(), key=lambda x: -x.frequency)


# ── Analysis: proper names ──

def detect_proper_names(type_list: list) -> set:
    """Detect proper names: at least one medial occurrence with uppercase first char."""
    proper = set()
    for te in type_list:
        has_medial_upper = any(
            occ.form_with_case[0].isupper() and occ.position_type == "sentence_medial"
            for occ in te.occurrences
        )
        if has_medial_upper:
            proper.add(te.folded)
            te.is_proper_name = True
    return proper


# ── XLSX styling ──

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(size=10)
LINK_FONT = Font(size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
GROUP_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TYPO_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
SUSP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PROPER_FONT = Font(size=10, color="2F5496")

# Für Spalten, die Glabbis manuell befüllt: klar erkennbar vom Rest abgesetzt.
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
INPUT_HEADER_FILL = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")


def style_input_column(ws, col: int, nrows: int):
    for row in range(2, nrows + 2):
        ws.cell(row=row, column=col).fill = INPUT_FILL


def highlight_form_in_context(context: str, form: str) -> CellRichText:
    """Return CellRichText with the *form* substring in red+bold."""
    idx = context.lower().find(form.lower())
    if idx < 0:
        return CellRichText(TextBlock(InlineFont(), context))

    before = context[:idx]
    match_str = context[idx:idx + len(form)]
    after = context[idx + len(form):]

    parts = []
    if before:
        parts.append(TextBlock(InlineFont(), before))
    parts.append(TextBlock(InlineFont(color="FF0000", b=True), match_str))
    if after:
        parts.append(TextBlock(InlineFont(), after))
    return CellRichText(*parts)


def _style_header(ws, ncols: int, input_cols: tuple = ()):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = INPUT_HEADER_FILL if col in input_cols else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        cell.border = THIN_BORDER


def _style_body(ws, nrows: int, ncols: int):
    for row in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _col_letter(i: int) -> str:
    """Column letter for 1-based index (supports >26)."""
    result = ""
    while i > 0:
        i, rem = divmod(i - 1, 26)
        result = chr(65 + rem) + result
    return result
