#!/usr/bin/env python3
"""Error-Classification: Korpus ↔ Twanksta-Dict, systematisch gruppiert
statt handgefrickelter Ausnahmelisten (Nachfolger-Werkzeug neben
delta_review.py, das unangetastet bleibt).

Drei Zielkategorien:
  1. Eigennamen / neue Wörter — Kandidaten für Wörterbuch-Erweiterungen
  2. Typos — echte Verschreiber (ED1 zu gut belegter Dict-Form)
  3. Phonologische Fehlerklassen — FST-verifiziert über den Ortho-Layer
     (fst/norm.regex → build/rules/<regel>.fst), kein Python-String-Diffing.
     Was keine einzelne Regel erklärt, landet ehrlich im Cluster-Fallback
     statt in einer erfundenen Kategorie.

YouTube-Übersetzungen (aus youtube_corpus_sentences.json) fließen aktiv in
die Klassifizierung von Bucket 1/2 ein: eine über mehrere Vorkommen stabile,
vom nächsten Wörterbuch-Lemma abweichende Übersetzung spricht für ein neues
Wort; eine mit dessen Glosse übereinstimmende Übersetzung spricht für Typo/
orthographische Variante.

Pipeline:
  1. Tokenize + exakte FST-Lookup gegen base.fst → known / unknown
  2. unknown gegen jede Ortho-Einzelregel (build/rules/*.fst) prüfen
     → genau 1 Regel: Bucket 3 (FST-Label)
     → mehrere Regeln ODER nur lenient.fst matcht: Cluster-Fallback
     → keine Regel matcht: weiter zu 3.
  3. ED1-Check gegen gut belegte Dict-Formen → Bucket 2 (Typo), mit
     Übersetzungs-Abgleich gegen die Glosse des Zielworts
  4. Rest: Bucket 1 (Eigennamen/neue Wörter), mit Übersetzungssignalen
  5. Output: error_classification.xlsx (3 Sheets) + stdout-Report
"""

import json
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl

from corpus_lookup import (
    REPO, CORPUS_PATH, BASE_FST, LENIENT_FST, RULES_DIR, ORTHO_RULES,
    fold, edit_distance, HAVE_RAPIDFUZZ,
    fst_lookup_batch, fst_lookup_types, load_dict_forms, check_precondition,
    tokenize_corpus, detect_proper_names,
    HEADER_FONT, HEADER_FILL, THIN_BORDER, TYPO_FILL, SUSP_FILL, PROPER_FONT,
    INPUT_FILL, style_input_column,
    highlight_form_in_context, _style_header, _style_body, _col_letter,
)

if HAVE_RAPIDFUZZ:
    from rapidfuzz import process as _rf_process
    from rapidfuzz.distance import Levenshtein as _Lev

OUT_DIR = REPO / "data/reports"
OUT_PATH = OUT_DIR / "error_classification.xlsx"

STOPWORDS = {
    "to", "a", "an", "the", "and", "of", "in", "on", "at", "is", "are", "be",
    "with", "for", "it", "this", "that", "or", "as", "by", "from", "not",
    "he", "she", "they", "we", "you", "i", "his", "her", "its", "their",
}


def _content_words(text: str) -> set[str]:
    words = {w.strip(".,!?;:\"'()").lower() for w in text.split()}
    return {w for w in words if len(w) >= 3 and w not in STOPWORDS}


def _all_case_forms(type_list: list) -> list[str]:
    """Every distinct surface spelling (all case variants) across all types
    — the same set fst_lookup_types builds internally, needed here so we can
    batch-query the per-rule transducers with the exact same strings."""
    forms = set()
    for te in type_list:
        forms.add(te.form)
        forms.add(te.form.lower())
        for occ in te.occurrences:
            forms.add(occ.form_with_case)
    return sorted(forms)


def classify_via_ortho_rules(unknown: list, all_forms: list) -> dict:
    """For every raw surface form, determine which named ortho-rule
    transducer(s) (build/rules/<name>.fst) accept it, plus whether the fully
    composed lenient.fst accepts it. Returns
    {form: {"rules": [rule_label, ...], "lenient": (lemma, analysis) | None}}"""
    result = {f: {"rules": [], "lenient": None} for f in all_forms}

    for rule_file, label in ORTHO_RULES:
        rule_path = RULES_DIR / f"{rule_file}.fst"
        if not rule_path.exists():
            print(f"  ⚠ {rule_path} fehlt (make all?) — übersprungen", file=sys.stderr)
            continue
        hits = fst_lookup_batch(all_forms, rule_path)
        for form, analyses in hits.items():
            lemma, analysis, tags = analyses[0]
            result[form]["rules"].append((label, lemma, analysis))

    lenient_hits = fst_lookup_batch(all_forms, LENIENT_FST)
    for form, analyses in lenient_hits.items():
        lemma, analysis, tags = analyses[0]
        result[form]["lenient"] = (lemma, analysis)

    return result


def build_dict_form_list(forms_dict: dict) -> list:
    return sorted(forms_dict["all_forms"])


def nearest_neighbor(ftok: str, forms_dict: dict, dict_form_list: list):
    if not HAVE_RAPIDFUZZ or len(ftok) < 3:
        return None, None, None, None
    m = _rf_process.extractOne(ftok, dict_form_list,
                                scorer=_Lev.normalized_distance)
    if not m:
        return None, None, None, None
    nn, nd, _idx = m
    return (nn,
            forms_dict["folded_to_orig"].get(nn, nn),
            forms_dict["folded_to_lemma"].get(nn, ["?"])[0],
            nd)


def run_pipeline() -> dict:
    t_start = time.time()
    report = {}

    print("Loading dictionary full-forms...", file=sys.stderr)
    forms_dict = load_dict_forms()
    prec = check_precondition(forms_dict)
    report["precondition"] = prec
    print(f"  {'✅ PASS' if prec['all_pass'] else '⚠ FAIL'} — "
          f"{prec['n_entries']} folded forms, {prec['n_lemmas']} lemmas",
          file=sys.stderr)

    print("Tokenizing corpus...", file=sys.stderr)
    type_list = tokenize_corpus()
    report["n_types_total"] = len(type_list)
    print(f"  {len(type_list)} types from corpus", file=sys.stderr)

    print("Detecting proper names (capitalization heuristic)...", file=sys.stderr)
    detect_proper_names(type_list)

    print("Exact FST lookup (base.fst only)...", file=sys.stderr)
    fst_known = fst_lookup_types(type_list, base_fst=BASE_FST, lenient_fst=None)
    known, unknown = [], []
    for td in type_list:
        if td.folded in fst_known:
            lemma, analysis = fst_known[td.folded]
            td.dict_lemma = lemma
            known.append(td)
        else:
            unknown.append(td)
    report["n_exact_hits"] = len(known)
    print(f"  {len(known)} exact hits, {len(unknown)} unmatched", file=sys.stderr)

    print("Checking ortho-layer rules (build/rules/*.fst)...", file=sys.stderr)
    unknown_forms = _all_case_forms(unknown)
    ortho_hits = classify_via_ortho_rules(unknown, unknown_forms)

    def _best_hit(td):
        """Merge per-form ortho-rule hits across all case variants of a type."""
        rule_labels = set()
        rule_info = None
        lenient_info = None
        forms = [td.form, td.form.lower()] + [o.form_with_case for o in td.occurrences]
        for f in forms:
            h = ortho_hits.get(f)
            if not h:
                continue
            for label, lemma, analysis in h["rules"]:
                rule_labels.add(label)
                if rule_info is None:
                    rule_info = (lemma, analysis)
            if h["lenient"] and lenient_info is None:
                lenient_info = h["lenient"]
        return sorted(rule_labels), rule_info, lenient_info

    bucket3_single = []      # exactly one ortho rule explains the form
    bucket3_fallback = []    # multiple rules or lenient-only match
    remaining = []           # no FST-based match at all

    for td in unknown:
        labels, rule_info, lenient_info = _best_hit(td)
        if len(labels) == 1 and rule_info:
            bucket3_single.append({
                "type": td, "rule": labels[0],
                "dict_lemma": rule_info[0], "analysis": rule_info[1],
            })
        elif len(labels) > 1:
            bucket3_fallback.append({
                "type": td, "reason": f"Mehrere Regeln: {', '.join(labels)}",
                "dict_lemma": rule_info[0] if rule_info else "",
            })
        elif lenient_info:
            bucket3_fallback.append({
                "type": td, "reason": "lenient.fst (rule combination), no single rule",
                "dict_lemma": lenient_info[0],
            })
        else:
            remaining.append(td)

    report["n_bucket3_single"] = len(bucket3_single)
    report["n_bucket3_fallback"] = len(bucket3_fallback)
    print(f"  {len(bucket3_single)} explained by exactly one ortho rule, "
          f"{len(bucket3_fallback)} cluster fallback, "
          f"{len(remaining)} still unmatched", file=sys.stderr)

    # Manche Formen stehen exakt so im rohen Twanksta-JSON (z.B. "pagaūwuns"),
    # sind aber nicht in base.fst/lenient.fst/den Ortho-Regeln kompiliert —
    # meist weil gen_lexc.py "↑"-Verweiseinträge (Partizip-/Flexionsformen,
    # die im Wörterbuch nur auf ein anderes Stichwort verweisen) überspringt.
    # Das ist ein FST-Build-Problem, kein neues Wort — verdient eine eigene,
    # klar getrennte Kategorie statt Bucket 1 zu verunreinigen.
    known_missing_fst = []
    truly_unknown = []
    for td in remaining:
        if td.folded in forms_dict["all_forms"]:
            lemma = forms_dict["folded_to_lemma"].get(td.folded, ["?"])[0]
            known_missing_fst.append({"type": td, "dict_lemma": lemma})
        else:
            truly_unknown.append(td)
    report["n_known_missing_fst"] = len(known_missing_fst)
    print(f"  {len(known_missing_fst)} already in dictionary JSON but missing "
          f"from compiled FST (likely ↑-reference entries skipped by "
          f"gen_lexc.py) — excluded from new-word bucket", file=sys.stderr)

    print("ED1 typo check against well-attested dict forms...", file=sys.stderr)
    dict_form_list = build_dict_form_list(forms_dict)
    type_freq = {te.folded: te.frequency for te in type_list}

    bucket2 = []
    bucket1 = []
    for td in truly_unknown:
        ftok = td.folded
        if td.frequency <= 2 and len(ftok) >= 3:
            nn_folded, nn_form, nn_lemma, nn_dist = nearest_neighbor(
                ftok, forms_dict, dict_form_list)
            if nn_folded and edit_distance(ftok, nn_folded) == 1:
                dict_freq = type_freq.get(nn_folded, 0)
                if dict_freq >= 5:
                    gloss = forms_dict["lemma_gloss"].get(fold(nn_lemma), [])
                    gloss_words = set()
                    for g in gloss:
                        gloss_words |= _content_words(g)
                    trans_words = set()
                    for hint in td.translation_hints():
                        trans_words |= _content_words(hint)
                    matches_gloss = bool(gloss_words & trans_words)
                    bucket2.append({
                        "type": td, "nn_form": nn_form, "nn_lemma": nn_lemma,
                        "matches_gloss": matches_gloss,
                        "gloss": ", ".join(gloss[:3]),
                    })
                    continue
        bucket1.append(td)

    report["n_bucket2"] = len(bucket2)
    print(f"  {len(bucket2)} typo candidates", file=sys.stderr)

    print("Translation signals for proper names/new words...", file=sys.stderr)
    bucket1_rows = []
    for td in bucket1:
        nn_folded, nn_form, nn_lemma, nn_dist = nearest_neighbor(
            td.folded, forms_dict, dict_form_list)
        hints = td.translation_hints()

        cap_cognate = False
        for hint in hints:
            words = hint.split()
            for i, w in enumerate(words):
                wc = w.strip(".,!?;:\"'()")
                if i == 0 or len(wc) < 2:
                    continue
                is_name_like = wc[0].isupper() and wc[1:].islower()
                if is_name_like and edit_distance(fold(wc), td.folded) <= 2:
                    cap_cognate = True

        gloss_overlap = False
        if nn_lemma:
            gloss = forms_dict["lemma_gloss"].get(fold(nn_lemma), [])
            gloss_words = set()
            for g in gloss:
                gloss_words |= _content_words(g)
            trans_words = set()
            for hint in hints:
                trans_words |= _content_words(hint)
            gloss_overlap = bool(gloss_words & trans_words)

        if td.is_proper_name or cap_cognate:
            signal = "Proper name (capitalization)" if td.is_proper_name else "Proper name (translation cognate)"
        elif hints and not gloss_overlap:
            signal = "New word (own meaning per translation)"
        elif gloss_overlap:
            signal = f"Note: translation matches gloss of \"{nn_lemma}\" — possibly a variant/typo"
        else:
            signal = "No translation signal"

        bucket1_rows.append({
            "type": td, "signal": signal,
            "nn_form": nn_form, "nn_lemma": nn_lemma,
        })

    report["n_bucket1"] = len(bucket1_rows)
    print(f"  {len(bucket1_rows)} proper names/new words", file=sys.stderr)

    report["bucket1"] = sorted(bucket1_rows, key=lambda r: -r["type"].frequency)
    report["bucket2"] = sorted(bucket2, key=lambda r: -r["type"].frequency)
    report["bucket3_single"] = sorted(bucket3_single, key=lambda r: (r["rule"], -r["type"].frequency))
    report["bucket3_fallback"] = sorted(bucket3_fallback, key=lambda r: -r["type"].frequency)
    report["known_missing_fst"] = sorted(known_missing_fst, key=lambda r: -r["type"].frequency)
    report["elapsed"] = time.time() - t_start
    return report


# ── XLSX output ──
#
# Alle Spalten-Header/Kategorien sind Englisch (Konvention aus
# delta_review.py — Glabbis' Review-Workflow ist Englisch, nur
# Code-Kommentare bleiben Deutsch). Kurzer Kontext-Ausschnitt statt
# ganzer Sätze (example_snippet), genau ein Übersetzungshinweis statt
# einer Liste, nur ein YouTube-Link, und am Ende leere, farblich
# abgesetzte "Glabbis:"-Spalten zum manuellen Ausfüllen.

def write_report(report: dict, path: Path):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Proper Names-New Words"
    headers1 = ["Form", "Freq", "Signal", "Translation",
                "Nearest Dict Form", "Context",
                "Glabbis: Lemma", "Glabbis: Comment", "YouTube"]
    for ci, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=ci, value=h)
    _style_header(ws1, len(headers1), input_cols=(7, 8))
    n1 = len(report["bucket1"])
    for ri, row in enumerate(report["bucket1"], 2):
        td = row["type"]
        ws1.cell(row=ri, column=1, value=td.form)
        ws1.cell(row=ri, column=2, value=td.frequency)
        ws1.cell(row=ri, column=3, value=row["signal"])
        ws1.cell(row=ri, column=4, value=td.translation_hint())
        nn = f"{row['nn_form']} ({row['nn_lemma']})" if row["nn_form"] else ""
        ws1.cell(row=ri, column=5, value=nn)
        ws1.cell(row=ri, column=6,
                 value=highlight_form_in_context(td.example_snippet(), td.form))
        ws1.cell(row=ri, column=9, value=td.top_links(1))
        if "Proper name" in row["signal"]:
            ws1.cell(row=ri, column=1).font = PROPER_FONT
    _style_body(ws1, n1, len(headers1))
    for col in (7, 8):
        style_input_column(ws1, col, n1)
    ws1.auto_filter.ref = f"A1:{_col_letter(len(headers1))}{n1 + 1}"
    ws1.freeze_panes = "A2"
    for i, w in enumerate([18, 7, 32, 26, 22, 42, 20, 26, 26], 1):
        ws1.column_dimensions[_col_letter(i)].width = w

    ws2 = wb.create_sheet("Typos")
    headers2 = ["Form", "Freq", "Dict Form", "Gloss matches translation?",
                "Gloss", "Context",
                "Glabbis: Confirmed?", "Glabbis: Comment", "YouTube"]
    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header(ws2, len(headers2), input_cols=(7, 8))
    n2 = len(report["bucket2"])
    for ri, row in enumerate(report["bucket2"], 2):
        td = row["type"]
        ws2.cell(row=ri, column=1, value=td.form)
        ws2.cell(row=ri, column=2, value=td.frequency)
        ws2.cell(row=ri, column=3, value=f"{row['nn_form']} ({row['nn_lemma']})")
        ws2.cell(row=ri, column=4, value="✓" if row["matches_gloss"] else "")
        ws2.cell(row=ri, column=5, value=row["gloss"])
        ws2.cell(row=ri, column=6,
                 value=highlight_form_in_context(td.example_snippet(), td.form))
        ws2.cell(row=ri, column=9, value=td.top_links(1))
        for col in (1, 2, 3, 4, 5, 6, 9):
            ws2.cell(row=ri, column=col).fill = TYPO_FILL
    _style_body(ws2, n2, len(headers2))
    for col in (7, 8):
        style_input_column(ws2, col, n2)
    ws2.auto_filter.ref = f"A1:{_col_letter(len(headers2))}{n2 + 1}"
    ws2.freeze_panes = "A2"
    for i, w in enumerate([18, 7, 24, 20, 30, 42, 16, 26, 26], 1):
        ws2.column_dimensions[_col_letter(i)].width = w

    ws3 = wb.create_sheet("Phonological Error Classes")
    headers3 = ["Form", "Freq", "Ortho Rule (FST)", "Dict Form-Lemma", "Context",
                "Glabbis: Comment", "YouTube"]
    for ci, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=ci, value=h)
    _style_header(ws3, len(headers3), input_cols=(6,))
    ri = 2
    grouped = defaultdict(list)
    for row in report["bucket3_single"]:
        grouped[row["rule"]].append(row)
    for rule_file, label in ORTHO_RULES:
        items = grouped.get(label, [])
        if not items:
            continue
        ws3.cell(row=ri, column=1,
                 value=f"▸ {label} ({len(items)} forms)")
        for col in range(2, len(headers3) + 1):
            ws3.cell(row=ri, column=col, value="")
        ws3.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(headers3))
        ri += 1
        for row in items:
            td = row["type"]
            ws3.cell(row=ri, column=1, value=td.form)
            ws3.cell(row=ri, column=2, value=td.frequency)
            ws3.cell(row=ri, column=3, value=label)
            ws3.cell(row=ri, column=4, value=row["dict_lemma"])
            ws3.cell(row=ri, column=5,
                     value=highlight_form_in_context(td.example_snippet(), td.form))
            ws3.cell(row=ri, column=7, value=td.top_links(1))
            ri += 1

    if report["bucket3_fallback"]:
        ws3.cell(row=ri, column=1,
                 value=f"▸ Unresolved / Cluster Fallback ({len(report['bucket3_fallback'])} forms)")
        for col in range(2, len(headers3) + 1):
            ws3.cell(row=ri, column=col, value="")
        ws3.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(headers3))
        ri += 1
        for row in report["bucket3_fallback"]:
            td = row["type"]
            ws3.cell(row=ri, column=1, value=td.form)
            ws3.cell(row=ri, column=2, value=td.frequency)
            ws3.cell(row=ri, column=3, value=row["reason"])
            ws3.cell(row=ri, column=4, value=row["dict_lemma"])
            ws3.cell(row=ri, column=5,
                     value=highlight_form_in_context(td.example_snippet(), td.form))
            ws3.cell(row=ri, column=7, value=td.top_links(1))
            for col in (1, 2, 3, 4, 5, 7):
                ws3.cell(row=ri, column=col).fill = SUSP_FILL
            ri += 1

    nrows3 = ri - 2
    _style_body(ws3, nrows3, len(headers3))
    style_input_column(ws3, 6, nrows3)
    ws3.auto_filter.ref = f"A1:{_col_letter(len(headers3))}{nrows3 + 1}"
    ws3.freeze_panes = "A2"
    for i, w in enumerate([18, 7, 30, 24, 42, 26, 26], 1):
        ws3.column_dimensions[_col_letter(i)].width = w

    # Nicht Teil der 3 inhaltlichen Kategorien: Formen, die exakt so im
    # rohen Twanksta-JSON stehen, aber nicht in base.fst/lenient.fst/den
    # Ortho-Regeln kompiliert sind (typ. "↑"-Verweiseinträge, die
    # gen_lexc.py überspringt). Kein neues Wort, kein Fehler — ein
    # FST-Build-Lücke. Eigenes Sheet, damit es Bucket 1 nicht verunreinigt.
    ws4 = wb.create_sheet("Known, Missing from FST")
    headers4 = ["Form", "Freq", "Dict Lemma", "Context",
                "Glabbis: Comment", "YouTube"]
    for ci, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=ci, value=h)
    _style_header(ws4, len(headers4), input_cols=(5,))
    n4 = len(report["known_missing_fst"])
    for ri4, row in enumerate(report["known_missing_fst"], 2):
        td = row["type"]
        ws4.cell(row=ri4, column=1, value=td.form)
        ws4.cell(row=ri4, column=2, value=td.frequency)
        ws4.cell(row=ri4, column=3, value=row["dict_lemma"])
        ws4.cell(row=ri4, column=4,
                 value=highlight_form_in_context(td.example_snippet(), td.form))
        ws4.cell(row=ri4, column=6, value=td.top_links(1))
    _style_body(ws4, n4, len(headers4))
    style_input_column(ws4, 5, n4)
    ws4.auto_filter.ref = f"A1:{_col_letter(len(headers4))}{n4 + 1}"
    ws4.freeze_panes = "A2"
    for i, w in enumerate([18, 7, 22, 42, 30, 26], 1):
        ws4.column_dimensions[_col_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  📄 {path}", file=sys.stderr)


def main():
    report = run_pipeline()
    write_report(report, OUT_PATH)

    r = report
    print(f"\n{'='*60}", file=sys.stderr)
    print("RUN REPORT", file=sys.stderr)
    print(f"{'='*60}", file=sys.stderr)
    print(f"  Precondition:            {'PASS' if r['precondition']['all_pass'] else 'FAIL'}", file=sys.stderr)
    print(f"  Types total:             {r['n_types_total']}", file=sys.stderr)
    print(f"  Exact hits (base.fst):   {r['n_exact_hits']}", file=sys.stderr)
    print(f"  Bucket 3 (1 rule):       {r['n_bucket3_single']}", file=sys.stderr)
    print(f"  Bucket 3 (fallback):     {r['n_bucket3_fallback']}", file=sys.stderr)
    print(f"  Bucket 2 (typos):        {r['n_bucket2']}", file=sys.stderr)
    print(f"  Bucket 1 (proper/new):   {r['n_bucket1']}", file=sys.stderr)
    print(f"  Known, missing FST:      {r['n_known_missing_fst']}", file=sys.stderr)
    print(f"  Elapsed:                 {r['elapsed']:.1f}s", file=sys.stderr)

    rule_counts = defaultdict(int)
    for row in r["bucket3_single"]:
        rule_counts[row["rule"]] += 1
    print(f"\n  Ortho-Regeln:", file=sys.stderr)
    for _, label in ORTHO_RULES:
        if rule_counts.get(label):
            print(f"    {label:<40} {rule_counts[label]:>5}", file=sys.stderr)

    summary = {
        "precondition_pass": r["precondition"]["all_pass"],
        "n_types_total": r["n_types_total"],
        "n_exact_hits": r["n_exact_hits"],
        "n_bucket1_eigennamen_neue_woerter": r["n_bucket1"],
        "n_bucket2_typos": r["n_bucket2"],
        "n_bucket3_ortho_single_rule": r["n_bucket3_single"],
        "n_bucket3_cluster_fallback": r["n_bucket3_fallback"],
        "n_known_missing_fst": r["n_known_missing_fst"],
        "elapsed_seconds": round(r["elapsed"], 1),
        "output": str(OUT_PATH),
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
