#!/usr/bin/env python3
"""Delta-Review: Korpus ↔ Twanksta-Dict (Handoff für Glabbis).

Pipeline:
  1. Precondition: verify dict is full-form
  2. Tokenize corpus → type list (frequency, context, source)
  3. Exact lookup each type → discard known
  4. Fuzzy lookup unknowns → classify as signature (A) or OOV (B)
  5. Cluster A by transformation signature
  6. Output delta_signatures.xlsx + delta_oov.xlsx + run report
"""

import json
import re
import unicodedata
import sys
import time
from collections import defaultdict, Counter
from pathlib import Path

try:
    from rapidfuzz.distance import Levenshtein as _Lev
    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Paths ──
REPO = Path(__file__).resolve().parents[2]
TWANKSTA_PATH = REPO / "data/external/twanksta_entries.json"
CORPUS_PATH = REPO / "corpus/youtube_corpus_sentences.json"
OUT_DIR = REPO / "data/reports"
OUT_SIG = OUT_DIR / "delta_signatures.xlsx"
OUT_OOV = OUT_DIR / "delta_oov.xlsx"

THRESHOLD = 0.25       # normalized Levenshtein
SKIP_VIDEOS = {"qLwBCWtMuH8"}

# ── Unicode helpers ──

def fold(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return unicodedata.normalize("NFC", s.lower())

def extract_prussian_tokens(text: str) -> list[str]:
    """Strip raw corpus text to clean Prussian tokens."""
    if re.match(r'^[A-Z]{2,5}:', text):
        if not text.startswith('PR:'):
            return []
        text = text[3:].lstrip()
    text = text.split("//")[0]
    text = re.sub(r'\[[^\]]*\]', '', text)
    text = re.sub(r'\([^)]*\)', '', text)
    tokens = []
    for tok in text.split():
        tok = tok.lstrip('=/')
        tok = tok.strip('.,!?;:()[]{}«»""\' \t')
        if tok and len(tok) >= 2 and tok.isalpha():
            tokens.append(tok)
    return tokens

# ── Levenshtein with backtrace ──

def edit_distance(a: str, b: str) -> int:
    if HAVE_RAPIDFUZZ:
        return _Lev.distance(a, b)
    m, n = len(a), len(b)
    if m < n:
        a, b = b, a
        m, n = n, m
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[0]
        dp[0] = i
        for j in range(1, n + 1):
            temp = dp[j]
            dp[j] = min(
                prev + (0 if a[i-1] == b[j-1] else 1),
                dp[j] + 1,
                dp[j-1] + 1,
            )
            prev = temp
    return dp[n]

def edit_ops(a: str, b: str) -> list[tuple[str, str, str, int, int]]:
    """Return list of (op, char_a, char_b, pos_a, pos_b)."""
    m, n = len(a), len(b)
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(m + 1): dp[i][0] = i
    for j in range(n + 1): dp[0][j] = j
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            cost = 0 if a[i-1] == b[j-1] else 1
            dp[i][j] = min(dp[i-1][j] + 1, dp[i][j-1] + 1, dp[i-1][j-1] + cost)
    ops = []
    i, j = m, n
    while i > 0 or j > 0:
        if i > 0 and j > 0 and a[i-1] == b[j-1]:
            ops.append(('eq', a[i-1], b[j-1], i-1, j-1))
            i -= 1; j -= 1
        elif i > 0 and j > 0 and dp[i][j] == dp[i-1][j-1] + 1:
            ops.append(('sub', a[i-1], b[j-1], i-1, j-1))
            i -= 1; j -= 1
        elif i > 0 and dp[i][j] == dp[i-1][j] + 1:
            ops.append(('del', a[i-1], '', i-1, j))
            i -= 1
        elif j > 0 and dp[i][j] == dp[i][j-1] + 1:
            ops.append(('ins', '', b[j-1], i, j-1))
            j -= 1
    ops.reverse()
    return ops

# ── Signature extraction ──

VOWEL_MAP = {'ā': 'a', 'a': 'ā', 'ē': 'e', 'e': 'ē', 'ī': 'i', 'i': 'ī',
             'ō': 'o', 'o': 'ō', 'ū': 'u', 'u': 'ū'}
CONSONANTS = set('bcdfghjklmnprstvwzšž')
VOWELS = set('aeiouāēīōū')

def ctx_label(ch: str, boundary: str = '#') -> str:
    if ch == '' or ch is None:
        return boundary
    if ch in CONSONANTS:
        return 'C'
    if ch in VOWELS:
        return 'V'
    return ch if ch.isalpha() else boundary

def derive_signature(ops: list, a: str, b: str) -> str:
    """Build a human-readable transformation signature from edit operations."""
    changes = [op for op in ops if op[0] != 'eq']
    if not changes:
        return "EXACT"

    lon = len(a)
    ltg = len(b)

    # Check gemination (single consonant ↔ double consonant)
    # Pattern: corpus has XX where dict has X, or vice versa
    if len(changes) == 2:
        c0, c1 = changes[0], changes[1]
        # double→single: DEL X + EQ X
        if (c0[0] == 'del' and c1[0] == 'eq'
                and c0[1] == c1[1] and c0[3] + 1 == c1[3]):
            ch = c0[1]
            lctx = ctx_label(a[c0[3]-1] if c0[3] > 0 else '')
            rctx = ctx_label(a[c0[3]+2] if c0[3]+2 < lon else '')
            return f"Gemination {ch}→{ch}{ch} (double→single) / {lctx}_{rctx}"
        # single→double: EQ X + INS X
        if (c0[0] == 'eq' and c1[0] == 'ins'
                and c0[1] == c1[2] and c0[4] + 1 == c1[4]):
            ch = c0[1]
            lctx = ctx_label(a[c0[3]-1] if c0[3] > 0 else '')
            rctx = ctx_label(a[c0[3]+1] if c0[3]+1 < lon else '')
            return f"Gemination {ch}→{ch}{ch} (single→double) / {lctx}_{rctx}"

    # Single operation cases
    if len(changes) == 1:
        op, ca, cb, pa, pb = changes[0]
        lctx_a = ctx_label(a[pa-1] if pa > 0 else '')
        rctx_a = ctx_label(a[pa+1] if pa + 1 < lon else '')
        if op == 'sub':
            if ca in VOWEL_MAP and cb == VOWEL_MAP.get(ca):
                return f"VLength {ca}→{cb} / {lctx_a}_{rctx_a}"
            return f"Sub {ca}→{cb} / {lctx_a}_{rctx_a}"
        if op == 'del':
            return f"Del {ca} / {lctx_a}_{rctx_a}"
        if op == 'ins':
            lctx_b = ctx_label(b[pb-1] if pb > 0 else '')
            rctx_b = ctx_label(b[pb+1] if pb + 1 < ltg else '')
            return f"Ins {cb} / {lctx_b}_{rctx_b}"

    # Multi-ops: try to identify vowel-length patterns (ā → a, a → ā)
    sub_ops = [op for op in changes if op[0] == 'sub']
    if sub_ops and all(op[1] in VOWEL_MAP and op[2] == VOWEL_MAP.get(op[1])
                       for op in sub_ops):
        details = ",".join(f"{op[1]}→{op[2]}" for op in sub_ops)
        return f"VLength multi ({details})"

    # Multi-ops: all substitutions
    if sub_ops and len(sub_ops) == len(changes):
        details = ",".join(f"{op[1]}→{op[2]}" for op in sub_ops)
        return f"Sub multi ({details})"

    # Composite fallback
    detail = ",".join(f"{op[0]}({op[1] or ''}→{op[2] or ''})" for op in changes)
    detail = detail[:60]
    return f"Composite ({detail})"


# ═══════════════════════════════════════════════════════════════
#  1. Precondition
# ═══════════════════════════════════════════════════════════════

def check_precondition(forms_dict: dict) -> dict:
    """Check that 10 inflected (non-lemma) forms can be found exactly."""
    lemma_samples = list(forms_dict["lemma_by_form"].items())
    inflected = [(f, lm) for f, lm in lemma_samples if f != lm and len(f) >= 3]
    test = inflected[:10]
    results = []
    for form, lemma in test:
        folded = fold(form)
        hit = folded in forms_dict["folded_to_lemma"]
        results.append({"form": form, "lemma": lemma, "folded": folded, "hit": hit})
    all_hit = all(r["hit"] for r in results)
    return {"tested": results, "all_pass": all_hit,
            "n_entries": len(forms_dict["all_forms"]),
            "n_original_forms": len(forms_dict["lemma_by_form"]),
            "n_lemmas": forms_dict["n_lemmas"]}


# ═══════════════════════════════════════════════════════════════
#  2. Load dictionary full forms
# ═══════════════════════════════════════════════════════════════

def load_dict_forms() -> dict:
    """Extract all inflected forms from Twanksta entries.

    Returns dict with:
      - all_forms: set of folded forms
      - forms_by_len: {len: [folded_form, ...]}
      - folded_to_lemma: folded→[lemma(s)]
      - folded_to_orig: folded→first original form
      - lemma_by_form: original_form→lemma
    """
    raw = json.loads(TWANKSTA_PATH.read_text(encoding="utf-8"))
    all_forms = set()
    forms_by_len = defaultdict(set)
    folded_to_lemma = defaultdict(list)
    folded_to_orig = {}
    lemma_by_form = {}

    for e in raw:
        word = e.get("word", "")
        if not word:
            continue
        fw = fold(word)
        all_forms.add(fw)
        forms_by_len[len(fw)].add(fw)
        folded_to_lemma[fw].append(word)
        folded_to_orig.setdefault(fw, word)
        lemma_by_form.setdefault(word, word)

        # declension tables
        for g in e.get("forms", {}).get("declension", []):
            for c in g.get("cases", []):
                for num in ("singular", "plural"):
                    f = c.get(num, "")
                    if f and " " not in f and "/" not in f:
                        ff = fold(f)
                        all_forms.add(ff)
                        forms_by_len[len(ff)].add(ff)
                        folded_to_lemma[ff].append(word)
                        folded_to_orig.setdefault(ff, f)
                        lemma_by_form.setdefault(f, word)

        # participles
        for p in e.get("forms", {}).get("participles", []):
            f = p.get("form", "")
            if f and " " not in f:
                ff = fold(f)
                all_forms.add(ff)
                forms_by_len[len(ff)].add(ff)
                folded_to_lemma[ff].append(word)
                folded_to_orig.setdefault(ff, f)
                lemma_by_form.setdefault(f, word)

        # verb forms
        for mood in ("indicative", "optative", "imperative", "subjunctive"):
            val = e.get("forms", {}).get(mood)
            if isinstance(val, list):
                for tense_entry in val:
                    if isinstance(tense_entry, dict):
                        for entry in tense_entry.get("forms", []):
                            f = entry.get("form", "")
                            if f and " " not in f and "/" not in f:
                                f = f.strip()
                                ff = fold(f)
                                all_forms.add(ff)
                                forms_by_len[len(ff)].add(ff)
                                folded_to_lemma[ff].append(word)
                                folded_to_orig.setdefault(ff, f)
                                lemma_by_form.setdefault(f, word)
                        f = tense_entry.get("form", "")
                        if f and " " not in f and "/" not in f:
                            f = f.strip()
                            ff = fold(f)
                            all_forms.add(ff)
                            forms_by_len[len(ff)].add(ff)
                            folded_to_lemma[ff].append(word)
                            folded_to_orig.setdefault(ff, f)
                            lemma_by_form.setdefault(f, word)
            elif isinstance(val, str):
                f = val.strip()
                if f and " " not in f and "/" not in f:
                    ff = fold(f)
                    all_forms.add(ff)
                    forms_by_len[len(ff)].add(ff)
                    folded_to_lemma[ff].append(word)
                    folded_to_orig.setdefault(ff, f)
                    lemma_by_form.setdefault(f, word)

    # Also build a mapping from folded → original (unfolded) forms
    # for computing edit ops on original strings
    folded_to_orig_form = {}
    for ff in all_forms:
        orig = folded_to_orig.get(ff)
        if orig:
            folded_to_orig_form[ff] = orig

    # Convert sets to sorted lists for deterministic iteration
    forms_by_len = {k: sorted(v) for k, v in forms_by_len.items()}

    # Count unique lemma words (not forms)
    n_lemmas = len(json.loads(TWANKSTA_PATH.read_text(encoding="utf-8")))

    return {
        "all_forms": all_forms,
        "forms_by_len": forms_by_len,
        "folded_to_lemma": dict(folded_to_lemma),
        "folded_to_orig": folded_to_orig,
        "folded_to_orig_form": folded_to_orig_form,
        "lemma_by_form": lemma_by_form,
        "n_lemmas": n_lemmas,
    }


# ═══════════════════════════════════════════════════════════════
#  3. Tokenize corpus → type list
# ═══════════════════════════════════════════════════════════════

def tokenize_corpus() -> list[dict]:
    """Tokenize corpus and build type list.

    Each entry: {form, folded, frequency, context_sentence, source_id}
    """
    raw = json.loads(CORPUS_PATH.read_text(encoding="utf-8"))
    type_data = {}  # folded → aggregate info

    for e in raw:
        sources = e.get("sources", [])
        if sources and all(s.get("video_id") in SKIP_VIDEOS for s in sources):
            continue

        text = e.get("text", "")
        for tok in extract_prussian_tokens(text):
            ftok = fold(tok)
            if ftok not in type_data:
                video_id = sources[0].get("video_id", "") if sources else ""
                type_data[ftok] = {
                    "form": tok,
                    "folded": ftok,
                    "frequency": 0,
                    "context_sentence": text[:120],
                    "source_id": video_id,
                }
            type_data[ftok]["frequency"] += 1

    return sorted(type_data.values(), key=lambda x: -x["frequency"])


# ═══════════════════════════════════════════════════════════════
#  4. Exact + fuzzy lookup
# ═══════════════════════════════════════════════════════════════

def lookup_exact(ftok: str, forms_dict: dict) -> list[str]:
    h = forms_dict["folded_to_lemma"].get(ftok)
    return h if h else []


def lookup_fuzzy(ftok: str, forms_dict: dict, threshold: float = THRESHOLD
                 ) -> list[tuple[str, float, str]]:
    """Find best fuzzy matches. Returns [(folded_dict_form, norm_distance, lemma)]."""
    if len(ftok) < 3:
        return []
    candidates = []
    ln = len(ftok)
    max_abs_diff = max(2, int(ln * threshold) + 1)
    forms_by_len = forms_dict["forms_by_len"]

    for diff in range(0, max_abs_diff + 1):
        for length in (ln - diff, ln + diff):
            if length <= 0 or length not in forms_by_len:
                continue
            for df in forms_by_len[length]:
                d = edit_distance(ftok, df)
                nd = d / max(ln, length)
                if nd <= threshold:
                    lemma = forms_dict["folded_to_lemma"].get(df, ["?"])[0]
                    candidates.append((df, nd, lemma))
        if len(candidates) > 100:
            break

    candidates.sort(key=lambda x: x[1])
    return candidates[:5]


# ═══════════════════════════════════════════════════════════════
#  5. Main pipeline
# ═══════════════════════════════════════════════════════════════

def run_pipeline() -> dict:
    t_start = time.time()
    report = {}

    # ── 1. Precondition ──
    print("Loading dictionary full-forms...", file=sys.stderr)
    forms_dict = load_dict_forms()
    prec = check_precondition(forms_dict)
    report["precondition"] = {
        "all_pass": prec["all_pass"],
        "n_entries": prec["n_entries"],
        "n_lemmas": prec["n_lemmas"],
        "tested": [
            {"form": r["form"], "lemma": r["lemma"], "hit": r["hit"]}
            for r in prec["tested"]
        ],
    }
    if not prec["all_pass"]:
        print("  ⚠ PRECONDITION FAILED — some forms not found exactly!", file=sys.stderr)
    else:
        print(f"  ✅ Precondition PASS — {prec['n_entries']} folded forms, {prec['n_lemmas']} lemmas",
              file=sys.stderr)

    # ── 2. Tokenize ──
    print("Tokenizing corpus...", file=sys.stderr)
    type_list = tokenize_corpus()
    report["n_types_total"] = len(type_list)
    report["n_corpus_entries"] = len(json.loads(CORPUS_PATH.read_text(encoding="utf-8")))
    print(f"  {len(type_list)} types from corpus", file=sys.stderr)

    # ── 3. Exact lookup ──
    print("Exact lookup...", file=sys.stderr)
    known = []
    unknown = []
    for td in type_list:
        hits = lookup_exact(td["folded"], forms_dict)
        if hits:
            td["dict_lemma"] = hits[0]
            known.append(td)
        else:
            unknown.append(td)
    report["n_exact_hits"] = len(known)
    print(f"  {len(known)} exact hits, {len(unknown)} unmatched", file=sys.stderr)

    # ── 4. Fuzzy lookup ──
    print("Fuzzy lookup...", file=sys.stderr)
    sig_candidates = []   # (type_data, best_match, norm_dist, lemma, ops, signature)
    oov_candidates = []   # type_data

    # Pre-pass: generate targeted variants (gemination, vowel length)
    # for fast matching before full edit distance search
    def _variant_pass(ftok: str, forms_dict: dict) -> list:
        """Try gemination and vowel-length variants first."""
        hits = []
        # gemination: double→single and single→double
        for m in re.finditer(r'(.)\1', ftok):
            alt = ftok[:m.start()] + m.group(1) + ftok[m.end():]
            if alt in forms_dict["all_forms"]:
                lemma = forms_dict["folded_to_lemma"].get(alt, ["?"])[0]
                nd = 1.0 / max(len(ftok), len(alt), 1)
                hits.append((alt, nd, lemma))
        for i in range(len(ftok)):
            if i+1 < len(ftok) and ftok[i] == ftok[i+1]:
                continue
            alt = ftok[:i+1] + ftok[i] + ftok[i+1:]
            if alt in forms_dict["all_forms"]:
                lemma = forms_dict["folded_to_lemma"].get(alt, ["?"])[0]
                nd = 1.0 / max(len(ftok), len(alt), 1)
                hits.append((alt, nd, lemma))
        # vowel length
        vpairs = {'a':'ā','ā':'a','e':'ē','ē':'e','i':'ī','ī':'i',
                  'o':'ō','ō':'o','u':'ū','ū':'u'}
        for i, ch in enumerate(ftok):
            if ch in vpairs:
                alt = ftok[:i] + vpairs[ch] + ftok[i+1:]
                if alt in forms_dict["all_forms"]:
                    lemma = forms_dict["folded_to_lemma"].get(alt, ["?"])[0]
                    nd = 1.0 / max(len(ftok), len(alt), 1)
                    hits.append((alt, nd, lemma))
        return hits

    def _add_sig_candidate(td, best_df, best_nd, best_lemma, forms_dict):
        """Compute edit ops and signature on ORIGINAL (unfolded) forms."""
        orig_corpus = unicodedata.normalize("NFC", td["form"])
        orig_dict = unicodedata.normalize("NFC",
            forms_dict["folded_to_orig"].get(best_df, best_df))
        # Compute edit ops on ORIGINAL forms to capture diacritic changes
        ops = edit_ops(orig_corpus, orig_dict)
        sig = derive_signature(ops, orig_corpus, orig_dict)
        # Compute distance on folded for normalization fairness
        f_corpus = fold(orig_corpus)
        f_dict = fold(orig_dict)
        d = edit_distance(f_corpus, f_dict)
        nd = d / max(len(f_corpus), len(f_dict), 1)
        return {
            "type": td,
            "best_folded": best_df,
            "best_form": orig_dict,
            "best_lemma": best_lemma,
            "norm_dist": nd,
            "ops": ops,
            "signature": sig,
            "edit_distance": d,
        }

    for td in unknown:
        ftok = td["folded"]
        # quick check: skip very short tokens (likely noise)
        if len(ftok) <= 2:
            oov_candidates.append({"type": td, "nearest_folded": None,
                                    "nearest_form": None, "nearest_lemma": None,
                                    "nearest_dist": None})
            continue

        # First try variant-based fast match
        variant_hits = _variant_pass(ftok, forms_dict)
        if variant_hits:
            best_df, best_nd, best_lemma = variant_hits[0]
            sig_candidates.append(
                _add_sig_candidate(td, best_df, best_nd, best_lemma, forms_dict))
            continue

        # Full edit distance search
        cands = lookup_fuzzy(ftok, forms_dict)
        if cands:
            best_df, best_nd, best_lemma = cands[0]
            sig_candidates.append(
                _add_sig_candidate(td, best_df, best_nd, best_lemma, forms_dict))
        else:
            # Nearest neighbor even if below threshold (for Tabelle B)
            best_df = None
            best_nd = None
            best_lemma = None
            best_orig = None
            # Find nearest neighbor regardless of threshold
            forms_by_len = forms_dict["forms_by_len"]
            nn = None
            nn_dist = float('inf')
            ln = len(ftok)
            for diff in range(0, min(5, max(ln - 1, 1))):
                for length in (ln - diff, ln + diff):
                    if length <= 0 or length not in forms_by_len:
                        continue
                    for df in forms_by_len[length]:
                        d = edit_distance(ftok, df)
                        if d < nn_dist:
                            nn_dist = d
                            nn = df
            if nn:
                nd = nn_dist / max(len(ftok), len(nn), 1)
                best_df = nn
                best_nd = nd
                best_lemma = forms_dict["folded_to_lemma"].get(nn, ["?"])[0]
                best_orig = forms_dict["folded_to_orig"].get(nn, nn)

            oov_candidates.append({
                "type": td,
                "nearest_folded": best_df,
                "nearest_form": best_orig,
                "nearest_lemma": best_lemma,
                "nearest_dist": best_nd,
            })

    report["n_signature_raw"] = len(sig_candidates)
    report["n_oov"] = len(oov_candidates)
    print(f"  {len(sig_candidates)} signature candidates, {len(oov_candidates)} OOV",
          file=sys.stderr)

    # ── 5. Filter known grammatical alternations (participles etc.) ──
    print("Filtering known grammatical alternations...", file=sys.stderr)
    PARTICIPLE_PAIRS = {
        ("wusis", "wuns"), ("wusi", "wuns"), ("wus", "wuns"),
        ("usis", "uns"), ("usi", "uns"),
        ("wunsi", "wuns"), ("unsi", "uns"),
    }
    def _is_grammatical_alt(sc: dict) -> bool:
        """Detect if the delta is a known grammatical (inflectional) alternation
        rather than an orthographic variant. Currently catches active participle
        suffix variants: -wuns/-wusis/-wusi/-wus."""
        orig_corpus = sc["type"]["form"]
        orig_dict = sc["best_form"]
        for suffix_c, suffix_d in PARTICIPLE_PAIRS:
            if orig_corpus.endswith(suffix_c) and orig_dict.endswith(suffix_d):
                return True
            if orig_dict.endswith(suffix_c) and orig_corpus.endswith(suffix_d):
                return True
        return False

    sig_candidates = [sc for sc in sig_candidates if not _is_grammatical_alt(sc)]
    n_filtered = report["n_signature_raw"] - len(sig_candidates)
    report["n_signature"] = len(sig_candidates)
    if n_filtered:
        print(f"  Filtered {n_filtered} grammatical alternations (participles)",
              file=sys.stderr)

    # ── 6. Cluster by signature ──
    print("Clustering by signature...", file=sys.stderr)
    sig_clusters = defaultdict(list)
    for sc in sig_candidates:
        sig_clusters[sc["signature"]].append(sc)

    # Sort clusters by total frequency
    cluster_list = []
    for sig, items in sig_clusters.items():
        total_freq = sum(it["type"]["frequency"] for it in items)
        n_types = len(items)
        # Pick up to 3 representative examples (highest freq)
        examples = sorted(items, key=lambda x: -x["type"]["frequency"])[:3]
        ex_data = []
        for ex in examples:
            td = ex["type"]
            ex_data.append({
                "corpus_form": td["form"],
                "dict_form": ex["best_form"],
                "dict_lemma": ex["best_lemma"],
                "context": td["context_sentence"],
                "source_id": td["source_id"],
                "frequency": td["frequency"],
                "edit_distance": ex["edit_distance"],
                "norm_dist": ex["norm_dist"],
            })
        cluster_list.append({
            "signature": sig,
            "total_frequency": total_freq,
            "n_types": n_types,
            "examples": ex_data,
        })
    cluster_list.sort(key=lambda x: -x["total_frequency"])

    # Assign stable IDs
    for i, cl in enumerate(cluster_list, 1):
        cl["sig_id"] = f"SIG-{i:03d}"

    report["n_clusters"] = len(cluster_list)

    # Sort OOV by frequency desc
    oov_list = sorted(oov_candidates, key=lambda x: -x["type"]["frequency"])
    report["oov_list"] = oov_list
    report["cluster_list"] = cluster_list
    report["threshold"] = THRESHOLD
    report["elapsed"] = time.time() - t_start
    report["forms_dict"] = forms_dict  # for downstream use

    return report


# ═══════════════════════════════════════════════════════════════
#  6. XLSX output
# ═══════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(size=10)
LINK_FONT = Font(size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_body(ws, nrows: int, ncols: int):
    for row in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_signatures_xlsx(cluster_list: list, path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Signaturen"
    headers = [
        "Signatur-ID", "Transformation", "Haeufigkeit (#Token)",
        "#Type", "Beispiel 1 (Korpus → Dict)", "Kontext 1", "Source 1",
        "Beispiel 2 (Korpus → Dict)", "Kontext 2", "Source 2",
        "Beispiel 3 (Korpus → Dict)", "Kontext 3", "Source 3",
        "Hypothese", "Glabbis: Urteil", "Glabbis: Erklaerung",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    for ri, cl in enumerate(cluster_list, 2):
        ws.cell(row=ri, column=1, value=cl["sig_id"])
        ws.cell(row=ri, column=2, value=cl["signature"])
        ws.cell(row=ri, column=3, value=cl["total_frequency"])
        ws.cell(row=ri, column=4, value=cl["n_types"])
        for ei in range(3):
            col_off = 5 + ei * 3
            if ei < len(cl["examples"]):
                ex = cl["examples"][ei]
                ws.cell(row=ri, column=col_off,
                        value=f"{ex['corpus_form']} → {ex['dict_form']} (Lemma: {ex['dict_lemma']})")
                ws.cell(row=ri, column=col_off + 1, value=ex["context"])
                ws.cell(row=ri, column=col_off + 2, value=ex["source_id"])
            else:
                ws.cell(row=ri, column=col_off, value="")
                ws.cell(row=ri, column=col_off + 1, value="")
                ws.cell(row=ri, column=col_off + 2, value="")
        ws.cell(row=ri, column=14,
                value="Konvention / Typo / fehlende Paradigmenzelle")
        ws.cell(row=ri, column=15, value="")  # Glabbis Urteil
        ws.cell(row=ri, column=16, value="")  # Glabbis Erklaerung

    _style_body(ws, len(cluster_list), len(headers))
    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(cluster_list) + 1}"
    ws.freeze_panes = "A2"

    # Column widths
    widths = [14, 30, 14, 8, 40, 45, 14, 40, 45, 14, 40, 45, 14, 35, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{i}"].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  📄 {path}", file=sys.stderr)


def write_oov_xlsx(oov_list: list, path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OOV"
    headers = [
        "Korpusform", "Frequenz", "Kontext + Source-ID",
        "Naechster Dict-Nachbar", "Naechste Distanz (norm.)",
        "Glabbis: Lemma", "Glabbis: Kommentar",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    for ri, oc in enumerate(oov_list, 2):
        td = oc["type"]
        ws.cell(row=ri, column=1, value=td["form"])
        ws.cell(row=ri, column=2, value=td["frequency"])
        ws.cell(row=ri, column=3, value=f"{td['context_sentence']} | src: {td['source_id']}")
        if oc.get("nearest_form") and oc.get("nearest_dist") is not None:
            ws.cell(row=ri, column=4,
                    value=f"{oc['nearest_form']} (Lemma: {oc['nearest_lemma']})")
            ws.cell(row=ri, column=5, value=round(oc["nearest_dist"], 3))
        else:
            ws.cell(row=ri, column=4, value="")
            ws.cell(row=ri, column=5, value="")
        ws.cell(row=ri, column=6, value="")
        ws.cell(row=ri, column=7, value="")

    _style_body(ws, len(oov_list), len(headers))
    ws.auto_filter.ref = f"A1:G{len(oov_list) + 1}"
    ws.freeze_panes = "A2"

    widths = [22, 10, 65, 35, 14, 20, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{i}"].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  📄 {path}", file=sys.stderr)


# ═══════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════

def main():
    report = run_pipeline()

    # Output XLSX
    print("\nWriting XLSX reports...", file=sys.stderr)
    write_signatures_xlsx(report["cluster_list"], OUT_SIG)
    write_oov_xlsx(report["oov_list"], OUT_OOV)

    # ── Run report ──
    r = report
    print(f"\n{'='*60}", file=sys.stderr)
    print("RUN REPORT", file=sys.stderr)
    print(f"{'='*60}", file=sys.stderr)
    print(f"  Precondition:      {'PASS' if r['precondition']['all_pass'] else 'FAIL'}", file=sys.stderr)
    print(f"  Dict folded forms: {r['precondition']['n_entries']}", file=sys.stderr)
    print(f"  Dict lemmas:       {r['precondition']['n_lemmas']}", file=sys.stderr)
    print(f"  Corpus entries:    {r['n_corpus_entries']}", file=sys.stderr)
    print(f"  Types total:       {r['n_types_total']}", file=sys.stderr)
    print(f"  Exact hits:        {r['n_exact_hits']}", file=sys.stderr)
    print(f"  Signature cluster: {r['n_clusters']}", file=sys.stderr)
    print(f"  Sig-instances:     {r['n_signature']}", file=sys.stderr)
    print(f"  OOV:               {r['n_oov']}", file=sys.stderr)
    print(f"  Threshold:         {r['threshold']}", file=sys.stderr)
    print(f"  Elapsed:           {r['elapsed']:.1f}s", file=sys.stderr)

    # Print top 20 signatures
    print(f"\n  Top 20 signature clusters:", file=sys.stderr)
    print(f"  {'Sig-ID':<10} {'Freq':>6} {'#Type':>6}  Transformation", file=sys.stderr)
    print(f"  {'-'*10} {'-'*6} {'-'*6}  {'-'*40}", file=sys.stderr)
    for cl in r["cluster_list"][:20]:
        print(f"  {cl['sig_id']:<10} {cl['total_frequency']:>6} {cl['n_types']:>6}  {cl['signature'][:50]}",
              file=sys.stderr)

    # Also print report as JSON to stdout for agent handoff
    summary = {
        "precondition_pass": r["precondition"]["all_pass"],
        "n_dict_folded_forms": r["precondition"]["n_entries"],
        "n_dict_lemmas": r["precondition"]["n_lemmas"],
        "n_types_total": r["n_types_total"],
        "n_exact_hits": r["n_exact_hits"],
        "n_signature_instances": r["n_signature"],
        "n_signature_clusters": r["n_clusters"],
        "n_oov": r["n_oov"],
        "threshold": r["threshold"],
        "elapsed_seconds": round(r["elapsed"], 1),
        "output_signatures": str(OUT_SIG),
        "output_oov": str(OUT_OOV),
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
